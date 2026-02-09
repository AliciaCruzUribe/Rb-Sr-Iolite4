# / Type: DRS
# / Name: RbSr_Neoma_AgeCorr
# / Authors: Cici Cruz-Uribe, Grant Craig, Hayward Melton, and Iolite Software
# / Description: A full Rb-Sr DRS that calculated correlated uncertainties and performs and age correction on the Rb/Sr ratio
# / References: None
# / Version: 1.0
# / Contact: support@iolite-software.com

from iolite import QtGui
from iolite import QtCore
from iolite.QtGui import QMessageBox, QFileDialog
from iolite.types import Result
import numpy as np
import statsmodels.api as sm
from scipy import stats
import math
from scipy import constants
from sklearn.decomposition import PCA
import os
import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not available. Excel export will fall back to text format.")

'''
Error correlation function for calculating rho between Sr87/Sr86 and Rb87/Sr86
'''
def Sr87Sr86_Rb87Sr86_error_corr(sel):
    """
    Calculate the error correlation (rho) between StdCorr_Sr87_Sr86_MBC and AgeCorr_Rb87_Sr86_MBC
    for a given selection.

    Parameters:
    sel: Selection object from iolite

    Returns:
    Result object containing the correlation coefficient
    """
    result = Result()

    try:
        StdCorr_Sr87_Sr86_MBC = data.timeSeries("StdCorr_Sr87_Sr86_MBC")
        AgeCorr_Rb87_Sr86_MBC = data.timeSeries("AgeCorr_Rb87_Sr86_MBC")
    except RuntimeError as e:
        print(f"Error accessing time series: {e}")
        return result

    # Get data for this selection
    sr_array = StdCorr_Sr87_Sr86_MBC.dataForSelection(sel)
    rb_array = AgeCorr_Rb87_Sr86_MBC.dataForSelection(sel)

    # Filter out NaN and negative values
    mask = np.logical_and(sr_array > 0, ~np.isnan(sr_array))
    mask = np.logical_and(mask, np.logical_and(rb_array > 0, ~np.isnan(rb_array)))

    sr_array = sr_array[mask]
    rb_array = rb_array[mask]

    # Need at least 2 points to calculate correlation
    if len(sr_array) >= 2 and len(rb_array) >= 2:
        # Calculate Pearson correlation coefficient
        corr_coef = np.corrcoef(sr_array, rb_array)[0, 1]
        result.setValue(corr_coef)
        print(f"Selection {sel.name}: rho = {corr_coef:.4f} (n={len(sr_array)})")
    else:
        print(f"Selection {sel.name}: insufficient data for correlation (n={len(sr_array)})")

    return result

def york_regression(x, y, x_err, y_err, rho=None, fix_intercept=False, fixed_int=None, fixed_int_err=None):
    """
    York et al. (2004) regression with anchored intercept support

    Parameters:
    x, y: arrays of data points
    x_err, y_err: arrays of 2SE uncertainties
    rho: array of error correlations
    fix_intercept: boolean, whether to anchor the intercept
    fixed_int: anchored intercept value (added as a data point at x=0)
    fixed_int_err: anchored intercept uncertainty (2SE)

    Returns:
    slope, intercept, slope_err (2SE), intercept_err (2SE), mswd, probability, n, dof, cov_ab (2SE scale)
    where n = total points used (including synthetic if anchored)
          dof = degrees of freedom
    """
    from scipy import stats

    n_original = len(x)

    # Handle anchored intercept by adding a synthetic data point
    if fix_intercept and fixed_int is not None:
        print(f"  Anchored intercept mode: adding synthetic point at (0, {fixed_int}) with uncertainty {fixed_int_err}")

        # Add synthetic data point at x=0, y=fixed_int
        x = np.append(x, 0.0)
        y = np.append(y, fixed_int)

        # Add uncertainties for the synthetic point
        # x uncertainty at intercept is essentially zero (very small)
        x_err = np.append(x_err, 1e-10)  # Very small x uncertainty

        # TESTING: If IsoplotR treats anchored intercept uncertainty as 1SE (not 2SE),
        # we need to double it to match the 2SE format of our other data
        # This way when York divides by 2, we get back to the original 1SE value
        y_err_synthetic = (fixed_int_err * 2.0) if fixed_int_err is not None else 1e-10
        y_err = np.append(y_err, y_err_synthetic)

        print(f"  TESTING: Treating fixed_int_err as 1SE, converting to 2SE for consistency")
        print(f"  Input uncertainty: {fixed_int_err} (assumed 1SE)")
        print(f"  Synthetic point y-uncertainty: {y_err_synthetic} (converted to 2SE)")

        # Add rho for synthetic point (zero correlation)
        if rho is None:
            rho = np.zeros(n_original + 1)
        elif np.isscalar(rho):
            # If rho is a scalar, broadcast to array for all original points, then add 0 for synthetic
            rho = np.append(np.full(n_original, rho), 0.0)
        else:
            # rho is already an array, just append 0 for synthetic point
            rho = np.append(rho, 0.0)

        # Update n to include synthetic point
        n = len(x)
        print(f"  Total points including synthetic: {n}")

        # Set flag to adjust DOF later (we didn't "lose" a parameter for intercept)
        using_anchored_intercept = True
    else:
        n = len(x)
        using_anchored_intercept = False

        if rho is None:
            rho = np.zeros(n)
        elif np.isscalar(rho):
            # If rho is a scalar, broadcast to array for all points
            rho = np.full(n, rho)

    # Convert 2SE to 1SE for the regression
    sigma_x = x_err / 2.0
    sigma_y = y_err / 2.0

    print(f"York regression inputs:")
    print(f"  n = {n} points (including {1 if using_anchored_intercept else 0} synthetic)")
    print(f"  x range: {x.min():.2f} to {x.max():.2f}")
    print(f"  y range: {y.min():.6f} to {y.max():.6f}")
    print(f"  sigma_x range: {sigma_x.min():.4f} to {sigma_x.max():.4f}")
    print(f"  sigma_y range: {sigma_y.min():.8f} to {sigma_y.max():.8f}")

    # Weights (standard York approach - no special treatment)
    omega_x = 1.0 / (sigma_x**2)
    omega_y = 1.0 / (sigma_y**2)

    # Initial guess for slope (weighted least squares)
    w_init = omega_y
    x_mean = np.average(x, weights=w_init)
    y_mean = np.average(y, weights=w_init)
    b = np.sum(w_init * (x - x_mean) * (y - y_mean)) / np.sum(w_init * (x - x_mean)**2)

    print(f"  Initial slope guess (WLS): {b:.6e}")

    # Iterative solution (standard York algorithm)
    max_iterations = 10000
    tolerance = 1e-15

    iteration_count = 0
    for iteration in range(max_iterations):
        iteration_count = iteration + 1
        b_old = b

        # York 2004, eq. 13 - Effective weights
        alpha_term = rho * np.sqrt(omega_x * omega_y)
        W = (omega_x * omega_y) / (omega_x + b**2 * omega_y - 2.0 * b * alpha_term)
        W = np.abs(W)

        # Weighted means
        sum_W = np.sum(W)
        x_bar = np.sum(W * x) / sum_W
        y_bar = np.sum(W * y) / sum_W

        U = x - x_bar
        V = y - y_bar

        # Beta values (York 2004)
        beta = W * (U / omega_y + b * V / omega_x - (b * U + V) * alpha_term / (omega_x * omega_y))

        # New slope (York 2004, eq. 14)
        denom = np.sum(W * beta * U)
        if abs(denom) > 1e-30:
            b = np.sum(W * beta * V) / denom
        else:
            print(f"  Convergence stopped: denominator near zero at iteration {iteration_count}")
            break

        # Check convergence with relative tolerance
        if abs(b - b_old) < tolerance * max(abs(b), 1.0):
            print(f"  Converged at iteration {iteration_count}, delta = {abs(b - b_old):.6e}")
            break

    if iteration_count >= max_iterations:
        print(f"  Warning: Max iterations ({max_iterations}) reached without full convergence")

    print(f"  Final slope after {iteration_count} iterations: {b:.10e}")

    # Final calculations
    a = y_bar - b * x_bar
    x_adj = x_bar + beta
    sigma_b = np.sqrt(1.0 / np.sum(W * (x_adj - x_bar)**2))
    sigma_a = np.sqrt(1.0 / sum_W + x_bar**2 / np.sum(W * (x_adj - x_bar)**2))

    # Goodness of fit
    y_fit = a + b * x
    S = np.sum(W * (y - y_fit)**2)

    # Adjust DOF for anchored intercept
    if using_anchored_intercept:
        # With anchored intercept, we have n points and fit 2 parameters
        # The synthetic point doesn't count as "real" DOF loss
        dof = n - 2
    else:
        # Normal case
        dof = n - 2

    if dof > 0:
        mswd = S / dof
        prob = 1.0 - stats.chi2.cdf(S, dof)
    else:
        mswd = 0.0
        prob = 0.0

    # Calculate covariance between slope and intercept
    # From York regression: cov(a, b) = -x_bar * var(b)
    # This is because a = y_bar - b * x_bar
    cov_ab = -x_bar * sigma_b**2

    print(f"York regression results:")
    print(f"  slope (b) = {b:.6e}")
    print(f"  intercept (a) = {a:.6f}")
    print(f"  sigma_b (1SE) = {sigma_b:.6e}")
    print(f"  sigma_a (1SE) = {sigma_a:.6e}")
    print(f"  cov(a,b) (1SE) = {cov_ab:.6e}")
    print(f"  MSWD = {mswd:.6f}")
    print(f"  DOF = {dof}")
    print(f"  n (total points used) = {n}")

    # Convert uncertainties back to 2SE
    sigma_b_2se = sigma_b * 2.0
    sigma_a_2se = sigma_a * 2.0
    # Covariance scales by 4 (2^2) when converting from 1SE to 2SE
    cov_ab_2se = cov_ab * 4.0

    return b, a, sigma_b_2se, sigma_a_2se, mswd, prob, n, dof, cov_ab_2se


def tls_regression_free_intercept(x, y):
    """
    Total Least Squares regression with free intercept (Model 2).
    Uses PCA to find best-fit line, ignoring analytical uncertainties.
    Jackknife resampling for uncertainty estimation.

    Based on IsoplotR's tls() function.

    Parameters:
    x, y: arrays of data points

    Returns:
    slope, intercept, slope_uncertainty (1SE), intercept_uncertainty (1SE), n, degrees_of_freedom
    """
    from sklearn.decomposition import PCA

    n = len(x)

    # Combine x and y into data matrix
    data_matrix = np.column_stack([x, y])

    # Perform PCA
    pca = PCA(n_components=1)
    pca.fit(data_matrix)

    # First principal component gives the direction of the line
    pc1 = pca.components_[0]
    slope = pc1[1] / pc1[0]

    # Intercept: line passes through centroid
    x_mean = np.mean(x)
    y_mean = np.mean(y)
    intercept = y_mean - slope * x_mean

    # Jackknife uncertainty estimation
    jack_slopes = np.zeros(n)
    jack_intercepts = np.zeros(n)

    for i in range(n):
        # Leave out point i
        x_jack = np.delete(x, i)
        y_jack = np.delete(y, i)

        # Fit PCA on remaining points
        data_jack = np.column_stack([x_jack, y_jack])
        pca_jack = PCA(n_components=1)
        pca_jack.fit(data_jack)

        pc1_jack = pca_jack.components_[0]
        slope_jack = pc1_jack[1] / pc1_jack[0]

        x_mean_jack = np.mean(x_jack)
        y_mean_jack = np.mean(y_jack)
        intercept_jack = y_mean_jack - slope_jack * x_mean_jack

        jack_slopes[i] = slope_jack
        jack_intercepts[i] = intercept_jack

    # Calculate jackknife standard errors
    slope_mean = np.mean(jack_slopes)
    intercept_mean = np.mean(jack_intercepts)

    slope_var = ((n - 1) / n) * np.sum((jack_slopes - slope_mean)**2)
    intercept_var = ((n - 1) / n) * np.sum((jack_intercepts - intercept_mean)**2)

    slope_err = np.sqrt(slope_var)  # 1SE
    intercept_err = np.sqrt(intercept_var)  # 1SE

    # Model 2 doesn't calculate MSWD/probability (no analytical uncertainties)
    degrees_of_freedom = n - 2

    return slope, intercept, slope_err, intercept_err, n, degrees_of_freedom


def tls_regression_fixed_intercept(x, y, intercept, intercept_uncert=0.0):
    """
    Total Least Squares regression with fixed intercept (Model 2).
    Uses orthogonal regression forced through fixed intercept.

    Based on IsoplotR's anchored.deming function in tls.R.

    Parameters:
    x, y: arrays of data points
    intercept: fixed y-intercept value
    intercept_uncert: uncertainty in fixed intercept (1-sigma, not used in TLS)

    Returns:
    slope, slope_uncertainty (1SE), n, degrees_of_freedom
    """
    n = len(x)

    def orthogonal_regression_fixed_intercept(x, y, fixed_int):
        """Find slope that minimizes orthogonal distance with fixed intercept"""
        y_shifted = y - fixed_int

        # For orthogonal regression through origin: slope = sum(x*y) / sum(x^2) for OLS
        # But for orthogonal (TLS), we need to minimize perpendicular distances
        # Using the analytical solution for orthogonal regression through origin

        # Sum of squares
        sum_xy = np.sum(x * y_shifted)
        sum_x2 = np.sum(x**2)
        sum_y2 = np.sum(y_shifted**2)

        # Orthogonal regression slope through origin
        # From eigenvalue problem: slope = (sum_y2 - sum_x2 + sqrt((sum_y2 - sum_x2)^2 + 4*sum_xy^2)) / (2*sum_xy)
        diff = sum_y2 - sum_x2
        discriminant = np.sqrt(diff**2 + 4*sum_xy**2)

        if sum_xy > 0:
            slope = (diff + discriminant) / (2 * sum_xy)
        elif sum_xy < 0:
            slope = (diff - discriminant) / (2 * sum_xy)
        else:
            slope = sum_y2 / sum_x2 if sum_x2 > 0 else 0

        return slope

    # Fit slope
    slope = orthogonal_regression_fixed_intercept(x, y, intercept)

    # Calculate uncertainty using IsoplotR's exact method
    # Based on anchored.deming function in tls.R
    y_shifted = y - intercept

    # Predicted y values
    y_pred = slope * x

    # Orthogonal residuals
    residuals = (y_shifted - y_pred) / np.sqrt(1 + slope**2)

    # Variance of residuals
    ve = np.var(residuals, ddof=1)

    # Calculate Hessian (second derivative of misfit function)
    def misfit_for_slope(b):
        """Misfit function for slope (intercept fixed)"""
        if isinstance(b, np.ndarray):
            b = b[0]
        resid = (y - intercept - b * x) / np.sqrt(1 + b**2)
        return np.sum(resid**2)

    eps = np.sqrt(np.finfo(float).eps)
    h = eps * max(abs(slope), 1.0)
    f_plus = misfit_for_slope(slope + h)
    f_center = misfit_for_slope(slope)
    f_minus = misfit_for_slope(slope - h)
    H = (f_plus - 2*f_center + f_minus) / (h**2)

    # Slope variance following IsoplotR: inverthess(H) * ve
    if abs(H) > 1e-10:
        slope_var = ve / H
    else:
        slope_var = ve / 1e-10  # Fallback

    slope_err = np.sqrt(abs(slope_var))  # 1SE

    # For fixed intercept, DOF = n - 1
    degrees_of_freedom = n - 1

    return slope, slope_err, n, degrees_of_freedom


def york_overdispersion_free_intercept(x, y, sx, sy, r=0):
    """
    York regression with overdispersion parameter, free intercept (Model 3).
    Adds geological dispersion variance to the regression.

    Based on IsoplotR's MLyork() function with model=3.

    Parameters:
    x, y: arrays of x and y data
    sx, sy: arrays of uncertainties in x and y (1-sigma)
    r: array or scalar of correlation coefficients between x and y errors (default 0)

    Returns:
    slope, intercept, slope_uncertainty (1SE), intercept_uncertainty (1SE),
    mswd, probability, dispersion, dispersion_uncertainty (1SE), degrees_of_freedom
    """
    from scipy.optimize import minimize

    x = np.asarray(x, dtype=np.float64)
    y = np.asarray(y, dtype=np.float64)
    sx = np.asarray(sx, dtype=np.float64)
    sy = np.asarray(sy, dtype=np.float64)

    if np.isscalar(r):
        r = np.full_like(x, float(r), dtype=np.float64)
    else:
        r = np.asarray(r, dtype=np.float64)

    n = len(x)

    # Get initial guess from standard York regression
    # york_regression takes 2SE, so convert 1SE to 2SE
    sx_2se = sx * 2.0
    sy_2se = sy * 2.0
    slope_init, intercept_init, slope_err_init_2se, intercept_err_init_2se, mswd_init, prob_init, n_init, dof_init, cov_init = york_regression(
        x, y, sx_2se, sy_2se, rho=r, fix_intercept=False
    )
    slope_err_init = slope_err_init_2se / 2.0  # Convert back to 1SE
    intercept_err_init = intercept_err_init_2se / 2.0

    # Initial dispersion guess based on MSWD
    if mswd_init > 1:
        disp_init = np.std(y - intercept_init - slope_init * x) * np.sqrt(max(0, mswd_init - 1))
    else:
        disp_init = 1e-6

    # Define negative log-likelihood
    def neg_log_likelihood(params):
        a = params[0]  # intercept
        b = params[1]  # slope
        lw = params[2]  # log(dispersion)
        w = np.exp(lw)  # dispersion (always positive)

        nll = 0.0

        for i in range(n):
            var_x = sx[i]**2
            var_y = sy[i]**2 + w**2  # wtype='a' (constant dispersion)
            cov_xy = r[i] * sx[i] * sy[i]

            det_E = var_x * var_y - cov_xy**2
            if det_E <= 0:
                return 1e10

            omega_xx = var_y / det_E
            omega_yy = var_x / det_E
            omega_xy = -cov_xy / det_E

            alpha = omega_xx + omega_yy * b**2 + 2 * omega_xy * b
            beta = (y[i] - a - b * x[i]) * (omega_xy + b * omega_yy)

            if alpha <= 0:
                return 1e10

            x_opt = x[i] + beta / alpha

            dx = x[i] - x_opt
            dy = y[i] - a - b * x_opt

            maha = (omega_xx * dx**2 + 2 * omega_xy * dx * dy + omega_yy * dy**2)
            nll += 0.5 * (np.log(det_E) + maha)

        return nll

    # Optimize
    init_params = [intercept_init, slope_init, np.log(max(disp_init, 1e-10))]
    result = minimize(neg_log_likelihood, init_params, method='BFGS',
                     options={'gtol': 1e-8, 'disp': False})

    intercept = result.x[0]
    slope = result.x[1]
    dispersion = np.exp(result.x[2])

    # Calculate uncertainties from 3x3 Hessian matrix
    try:
        n_params = 3
        hessian = np.zeros((n_params, n_params))
        params_opt = np.array([intercept, slope, result.x[2]])

        ndeps = 0.0031
        eps_vec = np.array([ndeps * max(abs(p), 1.0) for p in params_opt])

        f0 = neg_log_likelihood(params_opt)
        for i in range(n_params):
            for j in range(n_params):
                eps_i = eps_vec[i]
                eps_j = eps_vec[j]

                if i == j:
                    p_plus = params_opt.copy()
                    p_plus[i] += eps_i
                    p_minus = params_opt.copy()
                    p_minus[i] -= eps_i
                    hessian[i, i] = (neg_log_likelihood(p_plus) - 2*f0 + neg_log_likelihood(p_minus)) / (eps_i**2)
                else:
                    p_pp = params_opt.copy()
                    p_pp[i] += eps_i
                    p_pp[j] += eps_j
                    p_pm = params_opt.copy()
                    p_pm[i] += eps_i
                    p_pm[j] -= eps_j
                    p_mp = params_opt.copy()
                    p_mp[i] -= eps_i
                    p_mp[j] += eps_j
                    p_mm = params_opt.copy()
                    p_mm[i] -= eps_i
                    p_mm[j] -= eps_j
                    hessian[i, j] = (neg_log_likelihood(p_pp) - neg_log_likelihood(p_pm) -
                                    neg_log_likelihood(p_mp) + neg_log_likelihood(p_mm)) / (4 * eps_i * eps_j)

        cov_matrix = np.linalg.inv(hessian)

        intercept_err = np.sqrt(abs(cov_matrix[0, 0]))  # 1SE
        slope_err = np.sqrt(abs(cov_matrix[1, 1]))  # 1SE
        dispersion_err_log = np.sqrt(abs(cov_matrix[2, 2]))
        dispersion_err = dispersion * dispersion_err_log  # 1SE

    except (np.linalg.LinAlgError, ValueError, Exception):
        slope_err = slope_err_init
        intercept_err = intercept_err_init
        dispersion_err = 0.0

    # Calculate MSWD and probability with augmented uncertainties
    var_y_aug = sy**2 + dispersion**2
    chi_sq = 0.0
    for i in range(n):
        resid = y[i] - intercept - slope * x[i]
        chi_sq += resid**2 / var_y_aug[i]

    degrees_of_freedom = n - 2
    mswd = chi_sq / degrees_of_freedom if degrees_of_freedom > 0 else 0.0
    prob = 1.0 - stats.chi2.cdf(chi_sq, degrees_of_freedom) if degrees_of_freedom > 0 else 0.0

    return slope, intercept, slope_err, intercept_err, mswd, prob, dispersion, dispersion_err, degrees_of_freedom


def york_overdispersion_fixed_intercept(x, y, sx, sy, intercept, intercept_uncert=0.0, r=0):
    """
    York regression with overdispersion parameter, fixed intercept (Model 3).
    The intercept is truly fixed, and the anchor uncertainty IS the dispersion parameter.

    Based on IsoplotR's MLyork() function with model=3 and anchored intercept.

    Parameters:
    x, y: arrays of x and y data
    sx, sy: arrays of uncertainties in x and y (1-sigma)
    intercept: fixed y-intercept value
    intercept_uncert: uncertainty in fixed intercept (1-sigma) - THIS IS THE DISPERSION
    r: array or scalar of correlation coefficients between x and y errors (default 0)

    Returns:
    slope, slope_uncertainty (1SE), mswd, probability, dispersion, degrees_of_freedom
    """
    from scipy.optimize import minimize

    x = np.asarray(x, dtype=np.float64)
    y = np.asarray(y, dtype=np.float64)
    sx = np.asarray(sx, dtype=np.float64)
    sy = np.asarray(sy, dtype=np.float64)

    if np.isscalar(r):
        r = np.full_like(x, float(r), dtype=np.float64)
    else:
        r = np.asarray(r, dtype=np.float64)

    n = len(x)

    # The anchor uncertainty IS the dispersion (IsoplotR Model 3 behavior)
    dispersion = intercept_uncert if intercept_uncert > 0 else 0.0

    # Get initial slope guess from York regression
    # york_regression takes 2SE, so convert 1SE to 2SE
    sx_2se = sx * 2.0
    sy_2se = sy * 2.0
    slope_init, intercept_init, slope_err_init_2se, intercept_err_init_2se, mswd_init, prob_init, n_init, dof_init, cov_init = york_regression(
        x, y, sx_2se, sy_2se, rho=r,
        fix_intercept=True, fixed_int=intercept, fixed_int_err=intercept_uncert
    )

    # Define negative log-likelihood with fixed intercept and fixed dispersion
    def neg_log_likelihood_fixed(params):
        b = params[0]  # slope
        a = intercept  # fixed
        w = dispersion  # fixed at anchor uncertainty

        nll = 0.0

        for i in range(n):
            var_x = sx[i]**2
            var_y = sy[i]**2 + w**2  # wtype='a' (constant dispersion)
            cov_xy = r[i] * sx[i] * sy[i]

            det_E = var_x * var_y - cov_xy**2
            if det_E <= 0:
                return 1e10

            omega_xx = var_y / det_E
            omega_yy = var_x / det_E
            omega_xy = -cov_xy / det_E

            alpha = omega_xx + omega_yy * b**2 + 2 * omega_xy * b
            beta = (y[i] - a - b * x[i]) * (omega_xy + b * omega_yy)

            if alpha <= 0:
                return 1e10

            x_opt = x[i] + beta / alpha

            dx = x[i] - x_opt
            dy = y[i] - a - b * x_opt

            maha = (omega_xx * dx**2 + 2 * omega_xy * dx * dy + omega_yy * dy**2)
            nll += 0.5 * (np.log(det_E) + maha)

        return nll

    # Optimize slope only
    result = minimize(neg_log_likelihood_fixed, [slope_init], method='BFGS',
                     options={'gtol': 1e-8, 'disp': False})

    slope = result.x[0]

    # Calculate slope uncertainty from Hessian
    try:
        eps = 0.0031 * max(abs(slope), 1.0)
        f0 = neg_log_likelihood_fixed([slope])
        f_plus = neg_log_likelihood_fixed([slope + eps])
        f_minus = neg_log_likelihood_fixed([slope - eps])
        hessian_slope = (f_plus - 2*f0 + f_minus) / (eps**2)

        if hessian_slope > 0:
            slope_err = 1.0 / np.sqrt(hessian_slope)  # 1SE
        else:
            # Fallback
            slope_err = slope_init * 0.01
    except Exception:
        slope_err = slope_init * 0.01

    # Calculate MSWD with the fixed dispersion
    var_y_aug = sy**2 + dispersion**2
    chi_sq = 0.0
    for i in range(n):
        resid = y[i] - intercept - slope * x[i]
        chi_sq += resid**2 / var_y_aug[i]

    # DOF = n - 1 (only slope is fitted)
    degrees_of_freedom = n - 1
    mswd = chi_sq / degrees_of_freedom if degrees_of_freedom > 0 else 0.0
    prob = 1.0 - stats.chi2.cdf(chi_sq, degrees_of_freedom) if degrees_of_freedom > 0 else 0.0

    return slope, slope_err, mswd, prob, dispersion, degrees_of_freedom


def runDRS():

    drs.message("Starting Rb-Sr isotopes DRS...")
    drs.progress(0)

    # Get settings
    settings = drs.settings()
    print(settings)

    indexChannel = data.timeSeries(settings["IndexChannel"])
    rmName = settings["ReferenceMaterial"]
    maskOption = settings["Mask"]
    maskChannel = data.timeSeries(settings["MaskChannel"])
    cutoff = settings["MaskCutoff"]
    trim = settings["MaskTrim"]

    # Create debug messages for the settings being used
    IoLog.debug("indexChannelName = %s" % indexChannel.name)
    IoLog.debug(
        "Masking data  = True" if maskOption else "Masking data  = False")
    IoLog.debug("maskChannelName = %s" % maskChannel.name)
    IoLog.debug("maskCutoff = %f" % cutoff)
    IoLog.debug("maskTrim = %f" % trim)

    # Setup index time
    drs.message("Setting up index time...")
    drs.progress(5)
    drs.setIndexChannel(indexChannel)

    # Setup the mask
    if maskOption:
        drs.message("Making mask...")
        drs.progress(10)
        mask = drs.createMaskFromCutoff(maskChannel, cutoff, trim)
        data.createTimeSeries('mask', data.Intermediate,
                              indexChannel.time(), mask)
    else:
        mask = np.ones_like(indexChannel.data())
        data.createTimeSeries('mask', data.Intermediate,
                              indexChannel.time(), mask)

    # Interp onto index time and baseline subtract
    drs.message("Interpolating onto index time and baseline subtracting...")
    drs.progress(25)

    allInputChannels = data.timeSeriesList(data.Input)
    blGrp = None

    if len(data.selectionGroupList(data.Baseline)) > 1:
        IoLog.error("There are multiple baseline groups. Rb-Sr DRS cannot proceed...")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return
    elif len(data.selectionGroupList(data.Baseline)) < 1:
        IoLog.error("No baselines. Please select some baselines. Rb-Sr DRS cannot proceed...")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return
    else:
        blGrp = data.selectionGroupList(data.Baseline)[0]

    if len(blGrp.selections()) < 1:
        IoLog.error("No baseline selections. Please select some baselines. Rb-Sr DRS cannot proceed...")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    for counter, channel in enumerate(allInputChannels):
        drs.message("Baseline subtracting %s" % channel.name)
        drs.progress(25 + 50*counter/len(allInputChannels))

        drs.baselineSubtract(blGrp, [allInputChannels[counter]], mask, 25, 75)
        cps_ch = data.timeSeries(channel.name + '_CPS')
        input_ch = data.timeSeries(channel.name)
        cps_ch.setProperty(
            'Pre-shift mass', input_ch.property('Pre-shift mass'))
        cps_ch.setProperty('Post-shift mass',
                           input_ch.property('Post-shift mass'))

    drs.message("Calculating raw ratios...")
    drs.progress(50)

    # Declare the channels used in the calculations:
    Sr88_CPS = data.timeSeriesList(
        data.Intermediate, {'Element': 'Sr', 'Mass' : '88'})[0].data()
    Sr84F_CPS = data.timeSeriesList(
        data.Intermediate, {'Element': 'F', 'Mass' : '84'})[0].data()
    Sr86F_CPS = data.timeSeriesList(
        data.Intermediate, {'Element': 'F', 'Mass' : '86'})[0].data()
    Sr87F_CPS = data.timeSeriesList(
        data.Intermediate, {'Element': 'F', 'Mass' : '87'})[0].data()
    Sr88F_CPS = data.timeSeriesList(
        data.Intermediate, {'Element': 'F', 'Mass' : '88'})[0].data()

    if settings["UseRb85"]:
        try:
            Rb85_CPS = data.timeSeriesList(
                data.Intermediate, {'Element': 'Rb', 'Mass' : '85'})[0].data()
        except IndexError:
            IoLog.error("Could not find Rb85 channel. Was it measured?")
            drs.message("Error. See Messages")
            drs.progress(100)
            drs.finished()
            return

        Rb87c_CPS = Rb85_CPS * 0.38571

    else:
        Rb87_CPS = data.timeSeriesList(
            data.Intermediate, {'Element': 'Rb', 'Mass' : '87'})[0].data()

        Rb87c_CPS = Rb87_CPS - Sr88_CPS * Sr87F_CPS/Sr88F_CPS

    Sr87_Sr86_Raw = Sr87F_CPS/Sr86F_CPS
    Rb87_Sr86_Raw = Rb87c_CPS/Sr86F_CPS
    Sr87_Rb87_Raw = Sr87F_CPS/Rb87c_CPS
    Sr88_Sr86_Raw = Sr88F_CPS/Sr86F_CPS
    Sr84_Sr86_Raw = Sr84F_CPS/Sr86F_CPS
    Beta = np.log(8.37520938/(Sr88F_CPS/Sr86F_CPS))/np.log(87.9056/85.9093)
    Sr87m_Sr86m_Raw = Sr87F_CPS/Sr86F_CPS * pow(86.9089/85.9093, Beta)
    Rb87_Sr86m_Raw = Rb87c_CPS/Sr86F_CPS * pow(86.9089/85.9093, Beta)
    Sr84m_Sr86m_Raw = Sr84F_CPS/Sr86F_CPS * pow(83.9134/85.9093, Beta)

    # Gather up intermediate channels and add them as time series:
    int_channel_names = ['Rb87c_CPS', 'Sr87_Sr86_Raw', 'Rb87_Sr86_Raw', 'Sr87_Rb87_Raw', 'Sr88_Sr86_Raw', 'Sr84_Sr86_Raw', 'Beta', 'Sr87m_Sr86m_Raw', 'Rb87_Sr86m_Raw', 'Sr84m_Sr86m_Raw']
    int_channels = [Rb87c_CPS, Sr87_Sr86_Raw, Rb87_Sr86_Raw, Sr87_Rb87_Raw, Sr88_Sr86_Raw, Sr84_Sr86_Raw, Beta, Sr87m_Sr86m_Raw, Rb87_Sr86m_Raw, Sr84m_Sr86m_Raw]
    for name, channel in zip(int_channel_names, int_channels):
        data.createTimeSeries(name, data.Intermediate,
                              indexChannel.time(), channel)

    drs.message("Correcting ratios...")
    drs.progress(80)

    StdSpline_Rb87_Sr86 = data.spline(rmName, "Rb87_Sr86_Raw").data()
    try:
        StdValue_Rb87_Sr86 = data.referenceMaterialData(rmName)["87Rb/86Sr"].value()
    except KeyError:
        IoLog.error("There was no 87Rb/86Sr value in the " + rmName +
                      " datafile. Rb-Sr DRS cannot proceed.")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    print("StdSpline_Rb87_Sr86 mean = %f" % StdSpline_Rb87_Sr86.mean())
    print("StdValue_Rb87_Sr86 = %f" % StdValue_Rb87_Sr86)

    StdCorr_Rb87_Sr86 = (Rb87_Sr86_Raw) * StdValue_Rb87_Sr86 / StdSpline_Rb87_Sr86
    data.createTimeSeries('StdCorr_Rb87_Sr86', data.Output,
                          indexChannel.time(), StdCorr_Rb87_Sr86)

    StdSpline_Sr87_Sr86 = data.spline(rmName, "Sr87_Sr86_Raw").data()
    try:
        StdValue_Sr87_Sr86 = data.referenceMaterialData(rmName)["87Sr/86Sr"].value()
    except KeyError:
        IoLog.error("There was no 87Sr/86Sr value in the " + rmName +
                      " datafile. Rb-Sr DRS cannot proceed.")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    print("StdSpline_Sr87_Sr86 mean = %f" % StdSpline_Sr87_Sr86.mean())
    print("StdValue_Sr87_Sr86 = %f" % StdValue_Sr87_Sr86)

    StdCorr_Sr87_Sr86 = (Sr87_Sr86_Raw) * StdValue_Sr87_Sr86 / StdSpline_Sr87_Sr86
    data.createTimeSeries('StdCorr_Sr87_Sr86', data.Output,
                          indexChannel.time(), StdCorr_Sr87_Sr86)

    StdSpline_Rb87_Sr86m = data.spline(rmName, "Rb87_Sr86m_Raw").data()
    try:
        StdValue_Rb87_Sr86m = data.referenceMaterialData(rmName)["87Rb/86Sr"].value()
    except KeyError:
        IoLog.error("There was no 87Rb/86Sr value in the " + rmName +
                      " datafile. Rb-Sr DRS cannot proceed.")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    print("StdSpline_Rb87_Sr86m mean = %f" % StdSpline_Rb87_Sr86m.mean())
    print("StdValue_Rb87_Sr86m = %f" % StdValue_Rb87_Sr86m)

    StdCorr_Rb87_Sr86_MBC = (Rb87_Sr86m_Raw) * StdValue_Rb87_Sr86m / StdSpline_Rb87_Sr86m
    data.createTimeSeries('StdCorr_Rb87_Sr86_MBC', data.Output,
                          indexChannel.time(), StdCorr_Rb87_Sr86_MBC)

    StdSpline_Sr87m_Sr86m = data.spline(rmName, "Sr87m_Sr86m_Raw").data()
    try:
        StdValue_Sr87m_Sr86m = data.referenceMaterialData(rmName)["87Sr/86Sr"].value()
    except KeyError:
        IoLog.error("There was no 87Sr/86Sr value in the " + rmName +
                      " datafile. Rb-Sr DRS cannot proceed.")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    print("StdSpline_Sr87m_Sr86m mean = %f" % StdSpline_Sr87m_Sr86m.mean())
    print("StdValue_Sr87m_Sr86m = %f" % StdValue_Sr87m_Sr86m)

    StdCorr_Sr87_Sr86_MBC = (Sr87m_Sr86m_Raw) * StdValue_Sr87m_Sr86m / StdSpline_Sr87m_Sr86m
    data.createTimeSeries('StdCorr_Sr87_Sr86_MBC', data.Output,
                          indexChannel.time(), StdCorr_Sr87_Sr86_MBC)

    #print("Registering associated results...")
    #data.registerAssociatedResult("Selection_Age", calcSelectionAge)

    # Calculate 2SE for StdCorr_Rb87_Sr86_MBC
    print("\n=== Calculating 2SE for Rb87/Sr86 ===")

    rb_ts = data.timeSeries("StdCorr_Rb87_Sr86_MBC")
    time_array = rb_ts.time()
    rb_data = rb_ts.data()

    # Initialize array for uncertainties
    StdCorr_Rb87_Sr86_2SE = np.zeros_like(rb_data)

    # Get ALL selection groups
    all_groups = []
    all_groups.extend(data.selectionGroupList(data.ReferenceMaterial))
    all_groups.extend(data.selectionGroupList(data.Sample))

    for group in all_groups:
        for selection in group.selections():
            # Convert QDateTime to numeric timestamp
            sel_start = selection.startTime.toMSecsSinceEpoch() / 1000.0
            sel_end = selection.endTime.toMSecsSinceEpoch() / 1000.0

            # Find indices within this selection's time range
            mask = (time_array >= sel_start) & (time_array <= sel_end)

            if np.any(mask):
                # Get data for this selection
                rb_selection = rb_data[mask]
                rb_selection = rb_selection[np.isfinite(rb_selection)]

                n_points = len(rb_selection)

                if n_points >= 3:  # Need at least 3 points for outlier rejection
                    # Calculate initial mean and standard deviation
                    rb_mean_initial = np.mean(rb_selection)
                    rb_std_initial = np.std(rb_selection, ddof=1)

                    # Apply 2 SD outlier rejection (matching iolite's default)
                    outlier_mask = np.abs(rb_selection - rb_mean_initial) <= (2.0 * rb_std_initial)
                    rb_cleaned = rb_selection[outlier_mask]

                    n_cleaned = len(rb_cleaned)

                    if n_cleaned >= 2:
                        # Recalculate statistics on cleaned data
                        rb_std = np.std(rb_cleaned, ddof=1)
                        rb_uncert_2se = 2.0 * rb_std / np.sqrt(n_cleaned)

                        # Store same 2SE value for all points in this selection
                        StdCorr_Rb87_Sr86_2SE[mask] = rb_uncert_2se

            elif n_points == 2:
                # With only 2 points, can't do outlier rejection
                rb_std = np.std(rb_selection, ddof=1)
                rb_uncert_2se = 2.0 * rb_std / np.sqrt(n_points)
                StdCorr_Rb87_Sr86_2SE[mask] = rb_uncert_2se

    # Create output channel for uncertainties
    data.createTimeSeries('StdCorr_Rb87_Sr86_2SE', data.Output,
                      indexChannel.time(), StdCorr_Rb87_Sr86_2SE)

    print("Created channel: StdCorr_Rb87_Sr86_2SE")

    StdSpline_Sr87m_Sr86m = data.spline(rmName, "Sr87m_Sr86m_Raw").data()
    try:
        StdValue_Sr87m_Sr86m = data.referenceMaterialData(rmName)["87Sr/86Sr"].value()
    except KeyError:
        IoLog.error("There was no 87Sr/86Sr value in the " + rmName +
                      " datafile. Rb-Sr DRS cannot proceed.")
        drs.message("Error. See Messages")
        drs.progress(100)
        drs.finished()
        return

    print("StdSpline_Sr87m_Sr86m mean = %f" % StdSpline_Sr87m_Sr86m.mean())
    print("StdValue_Sr87m_Sr86m = %f" % StdValue_Sr87m_Sr86m)

    StdCorr_Sr87_Sr86_MBC = (Sr87m_Sr86m_Raw) * StdValue_Sr87m_Sr86m / StdSpline_Sr87m_Sr86m
    data.createTimeSeries('StdCorr_Sr87_Sr86_MBC', data.Output,
                          indexChannel.time(), StdCorr_Sr87_Sr86_MBC)

    # Calculate 2SE for StdCorr_Sr87_Sr86_MBC
    print("\n=== Calculating 2SE for Sr87/Sr86 ===")

    sr_ts = data.timeSeries("StdCorr_Sr87_Sr86_MBC")
    time_array = sr_ts.time()
    sr_data = sr_ts.data()

    # Initialize array for uncertainties
    StdCorr_Sr87_Sr86_2SE = np.zeros_like(sr_data)

    # Get ALL selection groups
    all_groups = []
    all_groups.extend(data.selectionGroupList(data.ReferenceMaterial))
    all_groups.extend(data.selectionGroupList(data.Sample))

    for group in all_groups:
        for selection in group.selections():
            # Convert QDateTime to numeric timestamp
            sel_start = selection.startTime.toMSecsSinceEpoch() / 1000.0
            sel_end = selection.endTime.toMSecsSinceEpoch() / 1000.0

            # Find indices within this selection's time range
            mask = (time_array >= sel_start) & (time_array <= sel_end)

            if np.any(mask):
                # Get data for this selection
                sr_selection = sr_data[mask]
                sr_selection = sr_selection[np.isfinite(sr_selection)]

                n_points = len(sr_selection)

                if n_points >= 3:  # Need at least 3 points for outlier rejection
                    # Calculate initial mean and standard deviation
                    sr_mean_initial = np.mean(sr_selection)
                    sr_std_initial = np.std(sr_selection, ddof=1)

                    # Apply 2 SD outlier rejection (matching iolite's default)
                    outlier_mask = np.abs(sr_selection - sr_mean_initial) <= (2.0 * sr_std_initial)
                    sr_cleaned = sr_selection[outlier_mask]

                    n_cleaned = len(sr_cleaned)

                    if n_cleaned >= 2:
                        # Recalculate statistics on cleaned data
                        sr_std = np.std(sr_cleaned, ddof=1)
                        sr_uncert_2se = 2.0 * sr_std / np.sqrt(n_cleaned)

                        # Store same 2SE value for all points in this selection
                        StdCorr_Sr87_Sr86_2SE[mask] = sr_uncert_2se

            elif n_points == 2:
                # With only 2 points, can't do outlier rejection
                sr_std = np.std(sr_selection, ddof=1)
                sr_uncert_2se = 2.0 * sr_std / np.sqrt(n_points)
                StdCorr_Sr87_Sr86_2SE[mask] = sr_uncert_2se

    # Create output channel for uncertainties
    data.createTimeSeries('StdCorr_Sr87_Sr86_2SE', data.Output,
                      indexChannel.time(), StdCorr_Sr87_Sr86_2SE)

    print("Created channel: StdCorr_Sr87_Sr86_2SE")

    # Calculate cycle-by-cycle uncertainties
    print("\n=== Calculating Cycle Uncertainties ===")
    try:
        calculateCycleUncertainties()
    except Exception as e:
        print(f"Warning: Could not calculate cycle uncertainties: {e}")
        import traceback
        traceback.print_exc()

    #Use a secondary mica to correct the Rb/Sr ratios
    # Get the saved setting value
    settings = drs.settings()  # Get the dictionary

    selected_group_name = settings.get("SecondaryRM", None)  # Safe way to get the value

    # Check if a group was selected
    if not selected_group_name:
        print("Error: No reference material group selected")
        drs.message("Error: No secondary RM selected")
        drs.progress(100)
        drs.finished()
        return

    print(f"Processing group: {selected_group_name}")

    # Get the selection group
    ref_group = data.selectionGroup(selected_group_name)

    if not ref_group:
        print(f"Error: Group '{selected_group_name}' not found")
        available_groups = [g.name for g in data.selectionGroupList(data.ReferenceMaterial)]
        print(f"Available RM groups: {available_groups}")
        drs.message(f"Error: Group '{selected_group_name}' not found")
        drs.progress(100)
        drs.finished()
        return

    selections = ref_group.selections()
    if not selections or len(selections) == 0:
        print(f"Error: No selections in group '{selected_group_name}'")
        drs.message(f"Error: No selections in {selected_group_name}")
        drs.progress(100)
        drs.finished()
        return

    print(f"Found {len(selections)} selections in group")

    # Collect data from individual selections
    # We need to use the mean values (results) for each selection
    Rb87_Sr86_sec = []
    Sr87_Sr86_sec = []

    print("\nCollecting data from selections:")

    Rb87_Sr86_values = []
    Sr87_Sr86_values = []
    Rb87_Sr86_uncerts = []
    Sr87_Sr86_uncerts = []
    rho_values = []  # Per-point error correlations

    # Check if we should use individual integrations
    use_individual_integrations = settings.get("UseIndividualIntegrations", False)

    if use_individual_integrations:
        print("Using individual integration data points (not selection means)")
    else:
        print("Using selection mean values")

    for selection in selections:
        if use_individual_integrations:
            # Get raw data from the selection (all points)
            try:
                rb_ts = data.timeSeries("StdCorr_Rb87_Sr86_MBC")
                sr_ts = data.timeSeries("StdCorr_Sr87_Sr86_MBC")
                rb_unc_ts = data.timeSeries("StdCorr_Rb87_Sr86_UNC")
                sr_unc_ts = data.timeSeries("StdCorr_Sr87_Sr86_UNC")

                # Get analytical rho for each integration
                try:
                    rho_ts = data.timeSeries("Rho_Analytical")
                    rho_data = rho_ts.dataForSelection(selection)
                except RuntimeError:
                    # If Rho_Analytical doesn't exist, use zeros
                    print("  Warning: Rho_Analytical channel not found, using rho=0")
                    rho_data = None

                rb_data = rb_ts.dataForSelection(selection)
                sr_data = sr_ts.dataForSelection(selection)
                rb_unc_data = rb_unc_ts.dataForSelection(selection)
                sr_unc_data = sr_unc_ts.dataForSelection(selection)

                # Filter out NaN and invalid values
                valid_mask = np.isfinite(rb_data) & np.isfinite(sr_data) & \
                            np.isfinite(rb_unc_data) & np.isfinite(sr_unc_data) & \
                            (rb_data > 0) & (sr_data > 0)

                # Also filter rho if available
                if rho_data is not None:
                    valid_mask = valid_mask & np.isfinite(rho_data)

                rb_valid = rb_data[valid_mask]
                sr_valid = sr_data[valid_mask]
                rb_unc_valid = rb_unc_data[valid_mask]
                sr_unc_valid = sr_unc_data[valid_mask]

                if rho_data is not None:
                    rho_valid = rho_data[valid_mask]
                else:
                    rho_valid = np.zeros(len(rb_valid))

                if len(rb_valid) == 0:
                    print("  " + selection.name + ": Skipping - No valid data points")
                    continue

                # Add all valid points from this selection
                Rb87_Sr86_sec.extend(rb_valid)
                Sr87_Sr86_sec.extend(sr_valid)
                Rb87_Sr86_uncerts.extend(rb_unc_valid)
                Sr87_Sr86_uncerts.extend(sr_unc_valid)
                rho_values.extend(rho_valid)

                print("  " + selection.name + ": Added " + str(len(rb_valid)) + " individual data points")

            except Exception as e:
                print("  " + selection.name + ": Error getting data - " + str(e))
                continue
        else:
        # Use iolite's native Results - this applies iolite's outlier rejection automatically
            try:
                rb_result = data.result(selection, data.timeSeries("StdCorr_Rb87_Sr86_MBC"))
                sr_result = data.result(selection, data.timeSeries("StdCorr_Sr87_Sr86_MBC"))
                rb_2se_result = data.result(selection, data.timeSeries("StdCorr_Rb87_Sr86_2SE"))
                sr_2se_result = data.result(selection, data.timeSeries("StdCorr_Sr87_Sr86_2SE"))

                rb_mean = rb_result.value()
                sr_mean = sr_result.value()
                rb_uncert = rb_2se_result.value()
                sr_uncert = sr_2se_result.value()

                # Check for None or NaN values
                if rb_mean is None or sr_mean is None or np.isnan(rb_mean) or np.isnan(sr_mean):
                    print("  " + selection.name + ": Skipping - NaN or None values")
                    continue

                if rb_uncert is None or sr_uncert is None:
                    print("  " + selection.name + ": Skipping - No uncertainty values")
                    continue

                # Get number of points from the selection
                n_points = len(data.timeSeries("StdCorr_Rb87_Sr86_MBC").dataForSelection(selection))

            except Exception as e:
                print("  " + selection.name + ": Error getting results - " + str(e))
                continue

            Rb87_Sr86_sec.append(rb_mean)
            Sr87_Sr86_sec.append(sr_mean)
            Rb87_Sr86_uncerts.append(rb_uncert)
            Sr87_Sr86_uncerts.append(sr_uncert)

            print("  " + selection.name + ": Rb/Sr=" + str(rb_mean) + " +/- " + str(rb_uncert) +
                  ", Sr/Sr=" + str(sr_mean) + " +/- " + str(sr_uncert) + " (" + str(n_points) + " points, 2SE)")

    if len(Rb87_Sr86_sec) < 2:
        print(f"\nError: Only found {len(Rb87_Sr86_sec)} data points, need at least 2 for regression")
        drs.message("Error: Not enough data for regression")
        drs.progress(100)
        drs.finished()
        return

    # Convert to numpy arrays
    Rb87_Sr86_sec = np.array(Rb87_Sr86_sec)
    Sr87_Sr86_sec = np.array(Sr87_Sr86_sec)
    Rb87_Sr86_uncert_2se = np.array(Rb87_Sr86_uncerts)  # Keep as 2SE
    Sr87_Sr86_uncert_2se = np.array(Sr87_Sr86_uncerts)  # Keep as 2SE

    # For Models 2 and 3, we need 1SE
    Rb87_Sr86_uncert_1se = Rb87_Sr86_uncert_2se / 2.0
    Sr87_Sr86_uncert_1se = Sr87_Sr86_uncert_2se / 2.0

    print("\nSuccessfully collected " + str(len(Rb87_Sr86_sec)) + " data points")
    print("Rb87/Sr86 range: " + str(Rb87_Sr86_sec.min()) + " to " + str(Rb87_Sr86_sec.max()))
    print("Sr87/Sr86 range: " + str(Sr87_Sr86_sec.min()) + " to " + str(Sr87_Sr86_sec.max()))
    print("Note: Uncertainties are 2SE")

    # Get settings
    settings = drs.settings()
    use_fixed_intercept = settings.get("UseFixedIntercept", True)
    fixed_intercept = settings.get("FixedIntercept", 0.7043)
    fixed_intercept_uncert_2se = settings.get("FixedInterceptUncert", 0.0002)  # 2SE from settings
    fixed_intercept_uncert_1se = fixed_intercept_uncert_2se / 2.0  # Convert to 1SE for Models 2 and 3
    regression_model = settings.get("RegressionModel", 0)  # 0=Model 1, 1=Model 2, 2=Model 3

    # Perform regression based on selected model
    from scipy import stats

    # Calculate error correlation if enabled (used for Models 1 and 3)
    use_corr_errors = settings.get("UseCorrelatedErrors", False)
    if use_corr_errors and len(Rb87_Sr86_sec) >= 2:
        if use_individual_integrations and len(rho_values) == len(Rb87_Sr86_sec):
            # Use per-point analytical rho values for individual integrations
            rho = np.array(rho_values)
            print(f"Using per-integration correlated errors: rho range = {rho.min():.4f} to {rho.max():.4f}, mean = {rho.mean():.4f}")
        else:
            # Use global correlation for selection means
            rho = np.corrcoef(Rb87_Sr86_sec, Sr87_Sr86_sec)[0, 1]
            print(f"Using correlated errors: rho = {rho:.4f}")
    else:
        rho = 0
        print("Using uncorrelated errors (r=0)")

    # Model names for display
    model_names = ["Model 1: York (MSWD scaling)", "Model 2: Total Least Squares", "Model 3: York + Overdispersion"]
    print(f"\n=== Using {model_names[regression_model]} ===")
    print(f"Intercept mode: {'Fixed' if use_fixed_intercept else 'Free'}")
    if use_fixed_intercept:
        print(f"Fixed intercept: {fixed_intercept}")

    try:
        # Variables to store regression results
        dispersion = None
        dispersion_err = None
        prob = None

        if regression_model == 0:
            # Model 1: York regression with MSWD scaling
            # Uses unified york_regression function with 2SE inputs/outputs
            slope, intercept, slope_uncert_2se, intercept_uncert_2se, mswd, prob, n_reg, degrees_of_freedom, cov_ab = york_regression(
                Rb87_Sr86_sec, Sr87_Sr86_sec,
                Rb87_Sr86_uncert_2se, Sr87_Sr86_uncert_2se,
                rho=rho if use_corr_errors else None,
                fix_intercept=use_fixed_intercept,
                fixed_int=fixed_intercept if use_fixed_intercept else None,
                fixed_int_err=fixed_intercept_uncert_1se if use_fixed_intercept else None  # York expects 1SE for anchor
            )
            # Convert slope uncertainty to 1SE for age calculation
            slope_uncert = slope_uncert_2se / 2.0
            intercept_uncert = intercept_uncert_2se / 2.0

        elif regression_model == 1:
            # Model 2: Total Least Squares (ignores analytical uncertainties)
            if use_fixed_intercept:
                slope, slope_uncert, n_points, degrees_of_freedom = tls_regression_fixed_intercept(
                    Rb87_Sr86_sec, Sr87_Sr86_sec,
                    fixed_intercept, fixed_intercept_uncert_1se
                )
                intercept = fixed_intercept
                intercept_uncert = 0.0
            else:
                slope, intercept, slope_uncert, intercept_uncert, n_points, degrees_of_freedom = tls_regression_free_intercept(
                    Rb87_Sr86_sec, Sr87_Sr86_sec
                )
            # TLS doesn't have MSWD
            mswd = None

        elif regression_model == 2:
            # Model 3: York + Overdispersion
            if use_fixed_intercept:
                slope, slope_uncert, mswd, prob, dispersion, degrees_of_freedom = york_overdispersion_fixed_intercept(
                    Rb87_Sr86_sec, Sr87_Sr86_sec,
                    Rb87_Sr86_uncert_1se, Sr87_Sr86_uncert_1se,
                    fixed_intercept, fixed_intercept_uncert_1se, r=rho
                )
                intercept = fixed_intercept
                intercept_uncert = 0.0
                dispersion_err = 0.0  # Fixed dispersion has no uncertainty
            else:
                slope, intercept, slope_uncert, intercept_uncert, mswd, prob, dispersion, dispersion_err, degrees_of_freedom = york_overdispersion_free_intercept(
                    Rb87_Sr86_sec, Sr87_Sr86_sec,
                    Rb87_Sr86_uncert_1se, Sr87_Sr86_uncert_1se, r=rho
                )

    except Exception as e:
        print(f"Could not calculate regression: {str(e)}")
        import traceback
        traceback.print_exc()
        drs.message(f"Error in {model_names[regression_model]} regression")
        drs.progress(100)
        drs.finished()
        return

    # For Model 2 (TLS), prob is None; for Models 1 and 3, prob was returned from the regression
    # No need to recalculate

    # Calculate age
    lambda_rb87 = 1.3972e-11
    age_ma = np.log(slope + 1) / lambda_rb87 / 1e6

    # Calculate age uncertainty based on model
    s_slope = slope_uncert  # This is 1SE for all models now
    alpha = 0.05

    if regression_model == 0:
        # Model 1: York with MSWD scaling
        if prob is not None and prob < alpha:
            # MSWD significantly > 1: apply overdispersion correction
            t_value = stats.t.ppf(1 - alpha/2, degrees_of_freedom)
            s_slope_corrected = s_slope * np.sqrt(mswd)
            print(f"  Note: p-value < {alpha}, applying overdispersion correction")
            print(f"  MSWD = {mswd}, multiplier = {np.sqrt(mswd)}")
        else:
            # MSWD ≈ 1 or scatter consistent with analytical uncertainties
            t_value = stats.norm.ppf(1 - alpha/2)
            s_slope_corrected = s_slope
        age_uncert_ma = t_value * s_slope_corrected / (slope + 1) / lambda_rb87 / 1e6

    elif regression_model == 1:
        # Model 2: TLS uses normal distribution
        t_value = stats.norm.ppf(1 - alpha/2)
        s_slope_corrected = s_slope
        age_uncert_ma = t_value * s_slope_corrected / (slope + 1) / lambda_rb87 / 1e6

    elif regression_model == 2:
        # Model 3: IsoplotR simply multiplies 1SE by 2 for 95% CI
        # slope_uncert is already 1SE from the regression
        age_uncert_ma = 2.0 * s_slope / (slope + 1) / lambda_rb87 / 1e6
        t_value = 2.0  # For display only

    # Print results
    print(f"\nRegression results for {selected_group_name}:")
    print(f"  Model: {model_names[regression_model]}")
    print(f"  Intercept mode: {'Fixed' if use_fixed_intercept else 'Free'}")
    print(f"  Slope: {slope} +/- {slope_uncert}")
    print(f"  Intercept: {intercept} +/- {intercept_uncert}")
    if mswd is not None:
        print(f"  MSWD: {mswd}")
    print(f"  Degrees of freedom: {degrees_of_freedom}")
    if prob is not None:
        print(f"  Probability: {prob}")
    if dispersion is not None:
        print(f"  Dispersion: {dispersion} +/- {dispersion_err if dispersion_err else 0}")
    print(f"  Age: {age_ma} +/- {age_uncert_ma} Ma")

    # Store regression results as intermediate channels (constant values)
    index_time = indexChannel.time()
    n_points = len(index_time)

    slope_array = np.full(n_points, slope)
    intercept_array = np.full(n_points, intercept)
    age_array = np.full(n_points, age_ma)
    uncert_array = np.full(n_points, age_uncert_ma)
    mswd_array = np.full(n_points, mswd if mswd is not None else np.nan)
    prob_array = np.full(n_points, prob if prob is not None else np.nan)

    data.createTimeSeries('SecondaryRM_Slope', data.Intermediate, index_time, slope_array)
    data.createTimeSeries('SecondaryRM_Intercept', data.Intermediate, index_time, intercept_array)
    data.createTimeSeries('SecondaryRM_Age_Ma', data.Intermediate, index_time, age_array)
    data.createTimeSeries('SecondaryRM_uncert_Ma', data.Intermediate, index_time, uncert_array)
    data.createTimeSeries('SecondaryRM_MSWD', data.Intermediate, index_time, mswd_array)
    data.createTimeSeries('SecondaryRM_Probability', data.Intermediate, index_time, prob_array)

    print("\nStored regression results as intermediate channels:")
    print("  - SecondaryRM_Slope")
    print("  - SecondaryRM_Intercept")
    print("  - SecondaryRM_Age_Ma")
    print("  - SecondaryRM_uncert_Ma")
    print("  - SecondaryRM_MSWD")
    print("  - SecondaryRM_Probability")

    # Apply age correction using the secondary RM
    secondary_rm_age_expected = settings.get("SecondaryRMAge", 0.0)
    correction_factor = 1.0
    rel_uncert_correction = 0.0

    if secondary_rm_age_expected > 0:
        print("\n=== Applying Age Correction ===")
        print(f"Expected Secondary RM Age: {secondary_rm_age_expected} Ma")
        print(f"Measured Secondary RM Age: {age_ma} Ma")

        # Calculate correction factor
        correction_factor = secondary_rm_age_expected / age_ma
        print(f"Correction factor: {correction_factor}")

        # Calculate relative uncertainty
        rel_uncert_correction = age_uncert_ma / age_ma

    # Store results in a dictionary for display and export
    results_dict = {
        'group_name': selected_group_name,
        'regression_model': regression_model,
        'use_fixed_intercept': use_fixed_intercept,
        'fixed_intercept_value': fixed_intercept if use_fixed_intercept else None,
        'use_corr_errors': use_corr_errors,
        'use_individual_integrations': use_individual_integrations,
        'age_ma': age_ma,
        'age_uncert_ma': age_uncert_ma,
        'slope': slope,
        'slope_uncert': slope_uncert,
        'intercept': intercept,
        'intercept_uncert': intercept_uncert,
        'mswd': mswd,
        'prob': prob,
        'dof': degrees_of_freedom,
        'n_points': len(Rb87_Sr86_sec),
        'dispersion': dispersion,
        'dispersion_err': dispersion_err,
        'expected_age': secondary_rm_age_expected,
        'correction_factor': correction_factor,
        'rel_uncert_correction': rel_uncert_correction,
    }

    # Save results to DRS settings so they persist and can be displayed in the widget
    drs.setSetting("LastRegressionResults", results_dict)

    # Show popup dialog with results
    showResultsDialog(results_dict)

    print("\nSecondary RM correction completed successfully!")

    if secondary_rm_age_expected > 0:

        # Apply correction to Rb87/Sr86 ratios
        StdCorr_Rb87_Sr86_MBC_data = data.timeSeries("StdCorr_Rb87_Sr86_MBC").data()
        AgeCorr_Rb87_Sr86_MBC = StdCorr_Rb87_Sr86_MBC_data / correction_factor

        # Create output channel
        data.createTimeSeries('AgeCorr_Rb87_Sr86_MBC', data.Output,
                              indexChannel.time(), AgeCorr_Rb87_Sr86_MBC)

        # Calculate 2SE with propagated uncertainty
        # Relative uncertainty in correction factor (from the measured secondary RM age)
        rel_uncert_correction = age_uncert_ma / age_ma
        print(f"Relative uncertainty in correction factor: {rel_uncert_correction * 100:.3f}%")

        # Calculate 2SE with propagated uncertainty for all samples
        rb_ts = data.timeSeries("StdCorr_Rb87_Sr86_MBC")
        time_array = rb_ts.time()
        rb_data = rb_ts.data()

        # Initialize array for uncertainties
        AgeCorr_Rb87_Sr86_2SE = np.zeros_like(rb_data)

        # Relative uncertainty in correction factor
        rel_uncert_correction = age_uncert_ma / age_ma

        # Get selections from ALL groups
        all_groups = []
        all_groups.extend(data.selectionGroupList(data.ReferenceMaterial))
        all_groups.extend(data.selectionGroupList(data.Sample))

        for group in all_groups:
            for selection in group.selections():
                sel_start = selection.startTime.toMSecsSinceEpoch() / 1000.0
                sel_end = selection.endTime.toMSecsSinceEpoch() / 1000.0
                mask = (time_array >= sel_start) & (time_array <= sel_end)

                if np.any(mask):
                    # Get data for this selection
                    rb_selection = rb_data[mask]
                    rb_selection = rb_selection[np.isfinite(rb_selection)]

                    n_points = len(rb_selection)

                    if n_points >= 3:  # Need at least 3 points for outlier rejection
                        # Calculate initial mean and standard deviation
                        rb_mean_initial = np.mean(rb_selection)
                        rb_std_initial = np.std(rb_selection, ddof=1)

                        # Apply 2 SD outlier rejection
                        outlier_mask = np.abs(rb_selection - rb_mean_initial) <= (2.0 * rb_std_initial)
                        rb_cleaned = rb_selection[outlier_mask]

                        n_cleaned = len(rb_cleaned)

                        if n_cleaned >= 2:
                            # Recalculate statistics on cleaned data
                            rb_mean = np.mean(rb_cleaned)
                            rb_std = np.std(rb_cleaned, ddof=1)
                            rb_uncert_2se = 2.0 * rb_std / np.sqrt(n_cleaned)

                            # Relative uncertainty
                            rel_uncert_rb = rb_uncert_2se / rb_mean

                            # Combined relative uncertainty (quadrature)
                            combined_rel_uncert = np.sqrt(rel_uncert_rb**2 + rel_uncert_correction**2)

                            # Apply to corrected mean
                            rb_corrected_mean = rb_mean / correction_factor
                            rb_corrected_uncert = rb_corrected_mean * combined_rel_uncert

                            # Store same uncertainty for all points in selection
                            AgeCorr_Rb87_Sr86_2SE[mask] = rb_corrected_uncert

                elif n_points == 2:
                    # With only 2 points, can't do outlier rejection
                    rb_mean = np.mean(rb_selection)
                    rb_std = np.std(rb_selection, ddof=1)
                    rb_uncert_2se = 2.0 * rb_std / np.sqrt(n_points)

                    # Relative uncertainty
                    rel_uncert_rb = rb_uncert_2se / rb_mean

                    # Combined relative uncertainty (quadrature)
                    combined_rel_uncert = np.sqrt(rel_uncert_rb**2 + rel_uncert_correction**2)

                    # Apply to corrected mean
                    rb_corrected_mean = rb_mean / correction_factor
                    rb_corrected_uncert = rb_corrected_mean * combined_rel_uncert

                    # Store same uncertainty for all points in selection
                    AgeCorr_Rb87_Sr86_2SE[mask] = rb_corrected_uncert

        # Create output channel
        data.createTimeSeries('AgeCorr_Rb87_Sr86_2SE', data.Output,
                    indexChannel.time(), AgeCorr_Rb87_Sr86_2SE)

        print("\nAge-corrected channels created:")
        print("  - AgeCorr_Rb87_Sr86_MBC")
        print("  - AgeCorr_Rb87_Sr86_2SE (with propagated uncertainty)")

        # Create AgeCorr_Rb87_Sr86_UNC (per-integration uncertainties)
        print("\n=== Creating Age-Corrected Per-Integration Uncertainties ===")

        # Get the StdCorr UNC channel
        try:
            StdCorr_Rb87_Sr86_UNC_ts = data.timeSeries("StdCorr_Rb87_Sr86_UNC")
            StdCorr_Rb87_Sr86_UNC_data = StdCorr_Rb87_Sr86_UNC_ts.data()

            # Calculate age-corrected per-integration uncertainties
            # For each integration point:
            # 1. Get the StdCorr_Rb87_Sr86_UNC value (absolute uncertainty)
            # 2. Get the StdCorr_Rb87_Sr86_MBC value to calculate relative uncertainty
            # 3. Combine in quadrature with age correction uncertainty
            # 4. Apply to age-corrected value

            StdCorr_Rb87_Sr86_MBC_data = data.timeSeries("StdCorr_Rb87_Sr86_MBC").data()

            # Relative uncertainty in correction factor (same for all points)
            rel_uncert_correction = age_uncert_ma / age_ma

            # Initialize array for age-corrected UNC
            AgeCorr_Rb87_Sr86_UNC = np.zeros_like(StdCorr_Rb87_Sr86_UNC_data)

            # For each data point
            filtered_count = 0
            for i in range(len(StdCorr_Rb87_Sr86_UNC_data)):
                if np.isfinite(StdCorr_Rb87_Sr86_MBC_data[i]) and StdCorr_Rb87_Sr86_MBC_data[i] > 0 and \
                   np.isfinite(StdCorr_Rb87_Sr86_UNC_data[i]) and StdCorr_Rb87_Sr86_UNC_data[i] > 0:
                    # Relative uncertainty from StdCorr UNC channel
                    rel_uncert_rb = StdCorr_Rb87_Sr86_UNC_data[i] / StdCorr_Rb87_Sr86_MBC_data[i]

                    # Skip points with unrealistically high relative uncertainties (>100% indicates bad data)
                    if rel_uncert_rb > 1.0:
                        AgeCorr_Rb87_Sr86_UNC[i] = 0.0
                        filtered_count += 1
                        continue

                    # Combined relative uncertainty (quadrature sum)
                    combined_rel_uncert = np.sqrt(rel_uncert_rb**2 + rel_uncert_correction**2)

                    # Age-corrected value
                    rb_corrected = StdCorr_Rb87_Sr86_MBC_data[i] / correction_factor

                    # Age-corrected uncertainty
                    uncert_value = rb_corrected * combined_rel_uncert

                    # Additional sanity check: uncertainty should be reasonable relative to value
                    # If uncertainty is > 10x the value, something is very wrong
                    if uncert_value > 10.0 * rb_corrected:
                        AgeCorr_Rb87_Sr86_UNC[i] = 0.0
                        filtered_count += 1
                    else:
                        AgeCorr_Rb87_Sr86_UNC[i] = uncert_value
                else:
                    AgeCorr_Rb87_Sr86_UNC[i] = 0.0

            # Create the output channel
            data.createTimeSeries('AgeCorr_Rb87_Sr86_UNC', data.Output,
                                indexChannel.time(), AgeCorr_Rb87_Sr86_UNC)

            print("Successfully created channel: AgeCorr_Rb87_Sr86_UNC")
            print(f"  Based on StdCorr_Rb87_Sr86_UNC with propagated age correction uncertainty")
            print(f"  Relative uncertainty from age correction: {rel_uncert_correction * 100:.3f}%")

            # Calculate statistics only on valid (non-zero) uncertainties
            valid_uncerts = AgeCorr_Rb87_Sr86_UNC[AgeCorr_Rb87_Sr86_UNC > 0]
            total_points = len(StdCorr_Rb87_Sr86_UNC_data)

            if len(valid_uncerts) > 0:
                print(f"  Total data points: {total_points}")
                print(f"  Valid uncertainty points: {len(valid_uncerts)} ({100*len(valid_uncerts)/total_points:.1f}%)")
                print(f"  Filtered points (bad data): {filtered_count}")
                print(f"  Uncertainty range: {np.min(valid_uncerts):.6f} to {np.max(valid_uncerts):.6f}")
                print(f"  Mean uncertainty: {np.mean(valid_uncerts):.6f}")
                print(f"  Median uncertainty: {np.median(valid_uncerts):.6f}")
            else:
                print(f"  Warning: No valid uncertainty values created")

        except Exception as e:
            print(f"Warning: Could not create AgeCorr_Rb87_Sr86_UNC channel: {e}")
            print("  StdCorr_Rb87_Sr86_UNC may not exist yet")

        # Register error correlation between Sr87/Sr86 and Rb87/Sr86
        print("\n=== Registering Error Correlations ===")
        data.registerAssociatedResult("87Sr/86Sr - 87Rb/86Sr Rho", Sr87Sr86_Rb87Sr86_error_corr)
        print("Registered: 87Sr/86Sr - 87Rb/86Sr Rho")

        # Calculate and create error correlation channel with per-selection rho values

        # Get time series objects
        sr_ts = data.timeSeries("StdCorr_Sr87_Sr86_MBC")
        rb_ts = data.timeSeries("AgeCorr_Rb87_Sr86_MBC")
        time_array = sr_ts.time()

        # Initialize rho array with zeros
        rho_array = np.zeros(len(time_array))

        print("\n=== Calculating Per-Selection Error Correlations ===")
        print("(Note: Values always calculated for display; only applied to age correction if box is checked)")

        # Get all selections from both RM and Sample groups
        all_groups = []
        all_groups.extend(data.selectionGroupList(data.ReferenceMaterial))
        all_groups.extend(data.selectionGroupList(data.Sample))

        # Calculate rho for each selection and populate the array
        for group in all_groups:
            for selection in group.selections():
                # Get data for this selection
                sr_selection = sr_ts.dataForSelection(selection)
                rb_selection = rb_ts.dataForSelection(selection)

                # Filter out NaN and negative values
                mask = np.logical_and(sr_selection > 0, ~np.isnan(sr_selection))
                mask = np.logical_and(mask, np.logical_and(rb_selection > 0, ~np.isnan(rb_selection)))

                sr_clean = sr_selection[mask]
                rb_clean = rb_selection[mask]

                # Calculate rho if we have enough data
                if len(sr_clean) >= 2 and len(rb_clean) >= 2:
                    rho_selection = np.corrcoef(sr_clean, rb_clean)[0, 1]
                    print(f"  {selection.name}: rho = {rho_selection:.4f} (n={len(sr_clean)})")

                    # Populate the time array for this selection
                    sel_start = selection.startTime.toMSecsSinceEpoch() / 1000.0
                    sel_end = selection.endTime.toMSecsSinceEpoch() / 1000.0
                    time_mask = (time_array >= sel_start) & (time_array <= sel_end)
                    rho_array[time_mask] = rho_selection
                else:
                    print(f"  {selection.name}: insufficient data for correlation (n={len(sr_clean)})")

        # Create output channel for error correlation
        data.createTimeSeries('Rho_Sr87Sr86_Rb87Sr86', data.Output, time_array, rho_array)
        print("  - Created Rho_Sr87Sr86_Rb87Sr86 channel with per-selection values")

    else:
        print("\nNo age correction applied (SecondaryRMAge = 0 or not set)")

    drs.message("Finished!")
    drs.progress(100)
    drs.finished()

def calculateCycleUncertainties():
    """
    Calculate cycle-by-cycle uncertainties for Rb87/Sr86 and Sr87/Sr86 ratios
    using amplifier noise and counting statistics.
    """
    # Physical constants
    V2CPS = 6.2414e+07         # Volts to CPS factor
    K = constants.k            # Boltzman constant
    e = constants.elementary_charge   # Elementary charge
    TK = 315.15                # Instrument temp in K
    R_Rb87 = 1.0e+11           # Rb87 resistor
    R_Sr87 = 1.0e+13           # Sr87 resistor
    R_Sr86 = 1.0e+13           # Sr86 resistor
    integTime = 1.             # Integration time in seconds

    # Get the CPS channels created by this DRS
    # These are the baseline-subtracted _CPS channels created in the main DRS
    try:
        # Try to get Rb87c first (the corrected Rb87 after Sr interference correction)
        Rb87_CPS_ts = data.timeSeries('Rb87c_CPS')
        Rb87_CPS = Rb87_CPS_ts.data()
        time_CPS = Rb87_CPS_ts.time()
        print("Using Rb87c_CPS channel")
    except RuntimeError:
        try:
            # Fall back to looking for channels by mass property
            Rb87_CPS = data.timeSeriesList(data.Intermediate, {'Element': 'Rb', 'Mass': '87'})[0].data()
            time_CPS = data.timeSeriesList(data.Intermediate, {'Element': 'Rb', 'Mass': '87'})[0].time()
            print("Using Rb87 channel from mass property")
        except (RuntimeError, IndexError):
            print("Could not find Rb87 CPS channel")
            return

    try:
        # Get Sr channels using the Element/Mass properties
        Sr86F_CPS_ts = data.timeSeriesList(data.Intermediate, {'Element': 'F', 'Mass': '86'})[0]
        Sr87F_CPS_ts = data.timeSeriesList(data.Intermediate, {'Element': 'F', 'Mass': '87'})[0]
        Sr86_CPS = Sr86F_CPS_ts.data()
        Sr87_CPS = Sr87F_CPS_ts.data()
        print("Using Sr86F and Sr87F channels")
    except (RuntimeError, IndexError):
        print("Could not get the Sr counts channels")
        return

    # Get the corrected ratio channels
    try:
        Sr87_Sr86 = data.timeSeries('StdCorr_Sr87_Sr86_MBC').data()
        Rb87_Sr86 = data.timeSeries('StdCorr_Rb87_Sr86_MBC').data()
        print("Using StdCorr MBC ratio channels")
    except RuntimeError:
        print("Could not get the MBC ratio channels")
        return

    # Check if data is in volts and convert to CPS if needed
    median_diff = np.abs(np.nanmedian(np.diff(Rb87_CPS)))
    unitsIsVolts = True  # Based on your system

    if unitsIsVolts:
        Rb87_CPS = Rb87_CPS * V2CPS
        Sr87_CPS = Sr87_CPS * V2CPS
        Sr86_CPS = Sr86_CPS * V2CPS
        print(f"Converted from volts to CPS (V2CPS = {V2CPS})")

    # Calculate point uncertainties for Rb87, Sr87, and Sr86
    # Amplifier noise on Rb87 (in CPS)
    s_amp_Rb87 = np.sqrt((4. * K * TK * integTime)/((e)**2 * R_Rb87))

    # Total noise on Rb87
    s_Rb87 = (s_amp_Rb87 / (Rb87_CPS * integTime))**2 + (1 / (Rb87_CPS * integTime))

    # Sr87
    # Amplifier noise on Sr87 (in CPS)
    s_amp_Sr87 = np.sqrt((4. * K * TK * integTime)/((e)**2 * R_Sr87))

    # Total noise on Sr87
    s_Sr87 = (s_amp_Sr87 / (Sr87_CPS * integTime))**2 + (1 / (Sr87_CPS * integTime))

    # Sr86
    # Amplifier noise on Sr86 (in CPS)
    s_amp_Sr86 = np.sqrt((4. * K * TK * integTime)/((e)**2 * R_Sr86))

    # Total noise on Sr86
    s_Sr86 = (s_amp_Sr86 / (Sr86_CPS * integTime))**2 + (1 / (Sr86_CPS * integTime))

    # 2 Sigma Uncertainties
    s_Rb87_86 = (np.sqrt(s_Rb87 + s_Sr86) * Rb87_Sr86 * 2.)
    s_Sr87_86 = (np.sqrt(s_Sr87 + s_Sr86) * Sr87_Sr86 * 2.)

    # Calculate analytical error correlation (rho) for each integration
    # Rho arises from the shared Sr86 denominator in both ratios
    # Formula: rho = var_rel(Sr86) / sqrt(var_rel(Rb87/Sr86) * var_rel(Sr87/Sr86))
    # where var_rel is the relative variance (sigma/value)^2
    # Since s_Rb87, s_Sr87, s_Sr86 are already relative variances:
    # rho = s_Sr86 / sqrt((s_Rb87 + s_Sr86) * (s_Sr87 + s_Sr86))

    denominator = np.sqrt((s_Rb87 + s_Sr86) * (s_Sr87 + s_Sr86))
    # Avoid division by zero
    rho_analytical = np.where(denominator > 0, s_Sr86 / denominator, 0.0)
    # Ensure rho is bounded between -1 and 1 (should always be positive for shared denominator)
    rho_analytical = np.clip(rho_analytical, -1.0, 1.0)

    # Create the output channels
    data.createTimeSeries("StdCorr_Sr87_Sr86_UNC", data.Output, time_CPS, s_Sr87_86)
    data.createTimeSeries("StdCorr_Rb87_Sr86_UNC", data.Output, time_CPS, s_Rb87_86)
    data.createTimeSeries("Rho_Analytical", data.Output, time_CPS, rho_analytical)
    print(f"Successfully created uncertainty channels:")
    print(f"  - StdCorr_Sr87_Sr86_UNC")
    print(f"  - StdCorr_Rb87_Sr86_UNC")
    print(f"  - Rho_Analytical (per-integration error correlation)")
    print(f"Boltzmann constant K = {K}")
    print(f"Rho_Analytical range: {np.nanmin(rho_analytical):.4f} to {np.nanmax(rho_analytical):.4f}")

# def updateSecondaryRMAge(rm_name, age_line_edit):
#     """
#     Update the age line edit when a different secondary RM is selected.
#     """
#     try:
#         rm_data = data.referenceMaterialData(rm_name)
#         if "Age" in rm_data:
#             default_age = float(rm_data["Age"].value())
#         else:
#             default_age = 0.0
#     except:
#         default_age = 0.0
#
#     age_line_edit.setText(str(default_age))
#     drs.setSetting("SecondaryRMAge", default_age)
def updateSecondaryRMAge(rm_name, age_line_edit, age_2se_line_edit=None,
                         intercept_line_edit=None, intercept_uncert_line_edit=None):
    """
    Update the age and uncertainty line edits when a different secondary RM is selected.
    """
    try:
        rm_data = data.referenceMaterialData(rm_name)
        if "Age" in rm_data:
            default_age = float(rm_data["Age"].value())
        else:
            default_age = 0.0
    except:
        default_age = 0.0

    age_line_edit.setText(str(default_age))
    drs.setSetting("SecondaryRMAge", default_age)

    # Update uncertainty if the line edit was provided
    if age_2se_line_edit is not None:
        try:
            rm_data = data.referenceMaterialData(rm_name)
            if "Age" in rm_data:
                default_age_2se = float(rm_data["Age"].uncertainty())
            else:
                default_age_2se = 0.0
        except:
            default_age_2se = 0.0

        age_2se_line_edit.setText(str(default_age_2se))
        drs.setSetting("SecondaryRMAge2SE", default_age_2se)

    # Update fixed intercept if the line edit was provided
    if intercept_line_edit is not None:
        try:
            rm_data = data.referenceMaterialData(rm_name)
            if "87Sr/86Sr initial" in rm_data:
                default_intercept = float(rm_data["87Sr/86Sr initial"].value())
            else:
                default_intercept = 0.7043
        except:
            default_intercept = 0.7043

        intercept_line_edit.setText(str(default_intercept))
        drs.setSetting("FixedIntercept", default_intercept)

    # Update fixed intercept uncertainty if the line edit was provided
    if intercept_uncert_line_edit is not None:
        try:
            rm_data = data.referenceMaterialData(rm_name)
            if "87Sr/86Sr initial" in rm_data:
                default_intercept_uncert = float(rm_data["87Sr/86Sr initial"].uncertainty())
            else:
                default_intercept_uncert = 0.0002
        except:
            default_intercept_uncert = 0.0002

        intercept_uncert_line_edit.setText(str(default_intercept_uncert))
        drs.setSetting("FixedInterceptUncert", default_intercept_uncert)


def createResultsTableModel(results_dict=None):
    """
    Create and populate a QStandardItemModel for displaying regression results.
    Returns the model with Parameter | Value | Uncertainty columns.
    """
    model = QtGui.QStandardItemModel()
    model.setHorizontalHeaderLabels(["Parameter", "Value", "Uncertainty"])

    if results_dict:
        updateResultsTableModel(model, results_dict)

    return model


def updateResultsTableModel(model, results_dict):
    """Update the table model data from results dictionary."""
    model.removeRows(0, model.rowCount())

    if not results_dict:
        return

    model_names = ["Model 1: York (MSWD)", "Model 2: TLS", "Model 3: York+Overdisp"]

    def addRow(param, value, uncert=""):
        items = [
            QtGui.QStandardItem(str(param)),
            QtGui.QStandardItem(str(value)),
            QtGui.QStandardItem(str(uncert))
        ]
        model.appendRow(items)

    # Settings section
    addRow("Secondary RM", results_dict.get('group_name', 'N/A'))
    addRow("Regression Model", model_names[results_dict.get('regression_model', 0)])
    addRow("Intercept Mode", 'Fixed' if results_dict.get('use_fixed_intercept', False) else 'Free')
    addRow("", "")  # Blank row separator

    # Main results
    addRow("Age (Ma, 95% CI)", f"{results_dict.get('age_ma', 0):.2f}", f"± {results_dict.get('age_uncert_ma', 0):.2f}")
    addRow("Slope", f"{results_dict.get('slope', 0):.6e}", f"± {results_dict.get('slope_uncert', 0):.2e}")
    addRow("Intercept (87Sr/86Sr)", f"{results_dict.get('intercept', 0):.6f}", f"± {results_dict.get('intercept_uncert', 0):.6f}")
    addRow("", "")  # Blank row separator

    # Statistics
    if results_dict.get('mswd') is not None:
        addRow("MSWD", f"{results_dict.get('mswd', 0):.4f}")
    if results_dict.get('prob') is not None:
        addRow("Probability", f"{results_dict.get('prob', 0):.4f}")
    addRow("Degrees of Freedom", str(results_dict.get('dof', 0)))
    addRow("Number of Points", str(results_dict.get('n_points', 0)))

    # Dispersion (Model 3)
    if results_dict.get('dispersion') is not None:
        disp_err = results_dict.get('dispersion_err', 0) or 0
        addRow("Dispersion", f"{results_dict.get('dispersion', 0):.6f}", f"± {disp_err:.6f}")

    # Age correction info
    if results_dict.get('expected_age', 0) > 0:
        addRow("", "")  # Blank row separator
        addRow("Expected Age (Ma)", f"{results_dict.get('expected_age', 0):.2f}")
        addRow("Correction Factor", f"{results_dict.get('correction_factor', 1):.6f}")


def getResultsTableData(results_dict):
    """
    Convert results dictionary to list of tuples for table/export.
    Returns list of (parameter, value, uncertainty) tuples.
    """
    if not results_dict:
        return []

    model_names = ["Model 1: York (MSWD scaling)", "Model 2: Total Least Squares", "Model 3: York + Overdispersion"]
    rows = []

    # Settings section
    rows.append(("SETTINGS", "", ""))
    rows.append(("Secondary RM", results_dict.get('group_name', 'N/A'), ""))
    rows.append(("Regression Model", model_names[results_dict.get('regression_model', 0)], ""))
    rows.append(("Intercept Mode", 'Fixed' if results_dict.get('use_fixed_intercept', False) else 'Free', ""))
    if results_dict.get('use_fixed_intercept', False):
        rows.append(("Fixed Intercept Value", f"{results_dict.get('fixed_intercept_value', 0):.6f}", ""))
    rows.append(("Correlated Errors", 'Yes' if results_dict.get('use_corr_errors', False) else 'No', ""))
    rows.append(("Individual Integrations", 'Yes' if results_dict.get('use_individual_integrations', False) else 'No', ""))
    rows.append(("", "", ""))

    # Main results
    rows.append(("RESULTS", "", ""))
    rows.append(("Age (Ma)", f"{results_dict.get('age_ma', 0):.4f}", f"{results_dict.get('age_uncert_ma', 0):.4f} (95% CI)"))
    rows.append(("Slope", f"{results_dict.get('slope', 0):.10e}", f"{results_dict.get('slope_uncert', 0):.10e} (1SE)"))
    rows.append(("Intercept (87Sr/86Sr)", f"{results_dict.get('intercept', 0):.8f}", f"{results_dict.get('intercept_uncert', 0):.8f} (1SE)"))
    rows.append(("", "", ""))

    # Statistics
    rows.append(("STATISTICS", "", ""))
    if results_dict.get('mswd') is not None:
        rows.append(("MSWD", f"{results_dict.get('mswd', 0):.6f}", ""))
    if results_dict.get('prob') is not None:
        rows.append(("Probability (p-value)", f"{results_dict.get('prob', 0):.6f}", ""))
    rows.append(("Degrees of Freedom", str(results_dict.get('dof', 0)), ""))
    rows.append(("Number of Points", str(results_dict.get('n_points', 0)), ""))

    if results_dict.get('dispersion') is not None:
        disp_err = results_dict.get('dispersion_err', 0) or 0
        rows.append(("Dispersion", f"{results_dict.get('dispersion', 0):.8f}", f"{disp_err:.8f} (1SE)"))
    rows.append(("", "", ""))

    # Age correction
    if results_dict.get('expected_age', 0) > 0:
        rows.append(("AGE CORRECTION", "", ""))
        rows.append(("Expected Age (Ma)", f"{results_dict.get('expected_age', 0):.4f}", ""))
        rows.append(("Measured Age (Ma)", f"{results_dict.get('age_ma', 0):.4f}", ""))
        rows.append(("Correction Factor", f"{results_dict.get('correction_factor', 1):.8f}", ""))
        rows.append(("Relative Uncertainty", f"{results_dict.get('rel_uncert_correction', 0) * 100:.4f}%", ""))

    return rows


def showResultsDialog(results_dict):
    """
    Display a popup dialog with the regression results.
    Called automatically after DRS completes.

    Parameters:
    results_dict: Dictionary containing all regression results
    """
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setWindowTitle("Secondary RM Isochron Results")

    # Build the results text
    model_names = ["Model 1: York (MSWD scaling)", "Model 2: Total Least Squares", "Model 3: York + Overdispersion"]

    text = f"<b>Secondary RM:</b> {results_dict.get('group_name', 'N/A')}<br>"
    text += f"<b>Model:</b> {model_names[results_dict.get('regression_model', 0)]}<br>"
    text += f"<b>Intercept mode:</b> {'Fixed' if results_dict.get('use_fixed_intercept', False) else 'Free'}<br><br>"

    text += f"<b>Age:</b> {results_dict.get('age_ma', 0):.2f} ± {results_dict.get('age_uncert_ma', 0):.2f} Ma (95% CI)<br><br>"

    text += f"<b>Slope:</b> {results_dict.get('slope', 0):.6e} ± {results_dict.get('slope_uncert', 0):.6e}<br>"
    text += f"<b>Intercept:</b> {results_dict.get('intercept', 0):.6f} ± {results_dict.get('intercept_uncert', 0):.6f}<br><br>"

    if results_dict.get('mswd') is not None:
        text += f"<b>MSWD:</b> {results_dict.get('mswd', 0):.4f}<br>"
    if results_dict.get('prob') is not None:
        text += f"<b>Probability:</b> {results_dict.get('prob', 0):.4f}<br>"
    text += f"<b>Degrees of freedom:</b> {results_dict.get('dof', 0)}<br>"
    text += f"<b>Number of points:</b> {results_dict.get('n_points', 0)}<br>"

    if results_dict.get('dispersion') is not None:
        disp_err = results_dict.get('dispersion_err', 0) or 0
        text += f"<br><b>Dispersion:</b> {results_dict.get('dispersion', 0):.6f} ± {disp_err:.6f}<br>"

    if results_dict.get('use_fixed_intercept', False):
        text += f"<br><b>Expected Age:</b> {results_dict.get('expected_age', 0):.2f} Ma<br>"
        text += f"<b>Correction factor:</b> {results_dict.get('correction_factor', 1):.6f}<br>"

    msg.setText(text)
    msg.setStandardButtons(QMessageBox.Ok)
    msg.exec_()


def formatResultsText(results_dict):
    """
    Format the results as plain text for display in the settings widget.

    Parameters:
    results_dict: Dictionary containing all regression results

    Returns:
    Formatted string
    """
    if not results_dict:
        return "No results yet. Run the DRS to see results."

    model_names = ["Model 1: York (MSWD scaling)", "Model 2: Total Least Squares", "Model 3: York + Overdispersion"]

    lines = []
    lines.append(f"Secondary RM: {results_dict.get('group_name', 'N/A')}")
    lines.append(f"Model: {model_names[results_dict.get('regression_model', 0)]}")
    lines.append(f"Intercept: {'Fixed' if results_dict.get('use_fixed_intercept', False) else 'Free'}")
    lines.append("")
    lines.append(f"Age: {results_dict.get('age_ma', 0):.2f} ± {results_dict.get('age_uncert_ma', 0):.2f} Ma")
    lines.append("")
    lines.append(f"Slope: {results_dict.get('slope', 0):.6e} ± {results_dict.get('slope_uncert', 0):.6e}")
    lines.append(f"Intercept: {results_dict.get('intercept', 0):.6f} ± {results_dict.get('intercept_uncert', 0):.6f}")

    if results_dict.get('mswd') is not None:
        lines.append(f"MSWD: {results_dict.get('mswd', 0):.4f}")
    if results_dict.get('prob') is not None:
        lines.append(f"Probability: {results_dict.get('prob', 0):.4f}")
    lines.append(f"DOF: {results_dict.get('dof', 0)}, n: {results_dict.get('n_points', 0)}")

    if results_dict.get('dispersion') is not None:
        disp_err = results_dict.get('dispersion_err', 0) or 0
        lines.append(f"Dispersion: {results_dict.get('dispersion', 0):.6f} ± {disp_err:.6f}")

    return "\n".join(lines)


def exportResults(results_dict):
    """
    Export regression results to an Excel file (.xlsx).
    Falls back to text file if openpyxl is not available.

    Parameters:
    results_dict: Dictionary containing all regression results
    """
    if not results_dict:
        QMessageBox.warning(None, "Export Error", "No results to export. Run the DRS first.")
        return

    # Try to get the active iolite data folder, fall back to home directory
    try:
        default_dir = data.dataPath()
        if not default_dir or not os.path.isdir(default_dir):
            default_dir = os.path.expanduser("~")
    except:
        default_dir = os.path.expanduser("~")

    default_name = f"RbSr_Isochron_Results_{results_dict.get('group_name', 'unknown')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if HAS_OPENPYXL:
        default_path = os.path.join(default_dir, default_name + ".xlsx")
        file_path_result = QFileDialog.getSaveFileName(
            None, "Export Results", default_path,
            "Excel Files (*.xlsx);;Text Files (*.txt);;All Files (*)"
        )
    else:
        default_path = os.path.join(default_dir, default_name + ".txt")
        file_path_result = QFileDialog.getSaveFileName(
            None, "Export Results", default_path,
            "Text Files (*.txt);;All Files (*)"
        )

    # QFileDialog.getSaveFileName returns a tuple (filename, filter) in PyQt
    if isinstance(file_path_result, tuple):
        file_path = file_path_result[0]
    else:
        file_path = file_path_result

    if not file_path:
        return

    # Check if user wants Excel or text format
    if file_path.lower().endswith('.xlsx') and HAS_OPENPYXL:
        exportToExcel(results_dict, file_path)
    else:
        exportToText(results_dict, file_path)


def exportToExcel(results_dict, file_path):
    """
    Export results to an Excel file with formatting.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Isochron Results"

        # Define styles
        header_font = Font(bold=True, size=12)
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title row
        ws['A1'] = "Rb-Sr Isochron Regression Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Exported: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:C2')

        # Column headers
        ws['A4'] = "Parameter"
        ws['B4'] = "Value"
        ws['C4'] = "Uncertainty"
        for col in ['A', 'B', 'C']:
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].border = thin_border

        # Get table data
        table_data = getResultsTableData(results_dict)

        # Write data starting at row 5
        row = 5
        for param, value, uncert in table_data:
            ws[f'A{row}'] = param
            ws[f'B{row}'] = value
            ws[f'C{row}'] = uncert

            # Style section headers
            if param in ["SETTINGS", "RESULTS", "STATISTICS", "AGE CORRECTION"]:
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = section_fill
                ws[f'B{row}'].fill = section_fill
                ws[f'C{row}'].fill = section_fill

            # Add borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = thin_border

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22

        # Save the workbook
        wb.save(file_path)

        QMessageBox.information(None, "Export Successful", f"Results exported to:\n{file_path}")

    except Exception as e:
        QMessageBox.critical(None, "Export Error", f"Failed to export to Excel:\n{str(e)}")
        import traceback
        traceback.print_exc()


def exportToText(results_dict, file_path):
    """
    Export results to a text file (fallback if openpyxl not available).
    """
    model_names = ["Model 1: York (MSWD scaling)", "Model 2: Total Least Squares", "Model 3: York + Overdispersion"]

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("Rb-Sr Isochron Regression Results\n")
            f.write(f"Exported: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")

            f.write("SETTINGS\n")
            f.write("-" * 40 + "\n")
            f.write(f"Secondary RM Group: {results_dict.get('group_name', 'N/A')}\n")
            f.write(f"Regression Model: {model_names[results_dict.get('regression_model', 0)]}\n")
            f.write(f"Intercept Mode: {'Fixed' if results_dict.get('use_fixed_intercept', False) else 'Free'}\n")
            if results_dict.get('use_fixed_intercept', False):
                f.write(f"Fixed Intercept Value: {results_dict.get('fixed_intercept_value', 'N/A')}\n")
            f.write(f"Correlated Errors: {'Yes' if results_dict.get('use_corr_errors', False) else 'No'}\n")
            f.write(f"Individual Integrations: {'Yes' if results_dict.get('use_individual_integrations', False) else 'No'}\n")
            f.write("\n")

            f.write("RESULTS\n")
            f.write("-" * 40 + "\n")
            f.write(f"Age: {results_dict.get('age_ma', 0):.4f} ± {results_dict.get('age_uncert_ma', 0):.4f} Ma (95% CI)\n")
            f.write(f"Slope: {results_dict.get('slope', 0):.10e} ± {results_dict.get('slope_uncert', 0):.10e} (1SE)\n")
            f.write(f"Intercept: {results_dict.get('intercept', 0):.8f} ± {results_dict.get('intercept_uncert', 0):.8f} (1SE)\n")
            f.write("\n")

            f.write("STATISTICS\n")
            f.write("-" * 40 + "\n")
            if results_dict.get('mswd') is not None:
                f.write(f"MSWD: {results_dict.get('mswd', 0):.6f}\n")
            if results_dict.get('prob') is not None:
                f.write(f"Probability (p-value): {results_dict.get('prob', 0):.6f}\n")
            f.write(f"Degrees of Freedom: {results_dict.get('dof', 0)}\n")
            f.write(f"Number of Data Points: {results_dict.get('n_points', 0)}\n")

            if results_dict.get('dispersion') is not None:
                f.write(f"Dispersion: {results_dict.get('dispersion', 0):.8f}\n")
                if results_dict.get('dispersion_err'):
                    f.write(f"Dispersion Uncertainty: {results_dict.get('dispersion_err', 0):.8f} (1SE)\n")
            f.write("\n")

            if results_dict.get('use_fixed_intercept', False) or results_dict.get('expected_age'):
                f.write("AGE CORRECTION\n")
                f.write("-" * 40 + "\n")
                f.write(f"Expected Age: {results_dict.get('expected_age', 0):.4f} Ma\n")
                f.write(f"Measured Age: {results_dict.get('age_ma', 0):.4f} Ma\n")
                f.write(f"Correction Factor: {results_dict.get('correction_factor', 1):.8f}\n")
                f.write(f"Relative Uncertainty: {results_dict.get('rel_uncert_correction', 0) * 100:.4f}%\n")

            f.write("\n" + "=" * 60 + "\n")

        QMessageBox.information(None, "Export Successful", f"Results exported to:\n{file_path}")

    except Exception as e:
        QMessageBox.critical(None, "Export Error", f"Failed to export results:\n{str(e)}")


def settingsWidget():
    """
    This function puts together a user interface to configure the DRS.
    Layout: Settings on left, Results table on right.

    It is important to have the last line of this function call:
    drs.setSettingsWidget(widget)
    """

    # Main widget with horizontal layout (left: settings, right: results)
    widget = QtGui.QWidget()
    mainLayout = QtGui.QHBoxLayout()
    widget.setLayout(mainLayout)

    # === LEFT SIDE: Settings ===
    leftWidget = QtGui.QWidget()
    formLayout = QtGui.QFormLayout()
    leftWidget.setLayout(formLayout)

    timeSeriesNames = data.timeSeriesNames(data.Input)
    defaultChannelName = ""
    if timeSeriesNames:
        defaultChannelName = timeSeriesNames[0]

    rmNames = data.selectionGroupNames(data.ReferenceMaterial)

    #Set default settings
    drs.setSetting("IndexChannel", defaultChannelName)
    drs.setSetting("ReferenceMaterial", "G_BCR2G")
    drs.setSetting("Mask", False)
    drs.setSetting("MaskChannel", defaultChannelName)
    drs.setSetting("MaskCutoff", 0.1)
    drs.setSetting("MaskTrim", 0.0)
    drs.setSetting("FixedIntercept", 0.7073)
    drs.setSetting("FixedInterceptUncert", 0.0002)  # 2SE uncertainty
    drs.setSetting("UseIndividualIntegrations", False)
    drs.setSetting("UseFixedIntercept", False)
    drs.setSetting("UseCorrelatedErrors", False)
    drs.setSetting("RegressionModel", 0)  # 0=Model 1 (York), 1=Model 2 (TLS), 2=Model 3 (York+Overdispersion)
    drs.setSetting("UseRb85", True)
    drs.setSetting("SecondaryRM", "F_Shap")

    # Refresh settings after all defaults are set
    settings = drs.settings()

    # --- Primary Reference Material Settings ---
    generalLabel = QtGui.QLabel("<b>Primary Reference Material Settings</b>")
    formLayout.addRow(generalLabel)

    indexComboBox = QtGui.QComboBox(leftWidget)
    indexComboBox.addItems(timeSeriesNames)
    indexComboBox.setCurrentText(settings["IndexChannel"])
    indexComboBox.currentTextChanged.connect(
        lambda t: drs.setSetting("IndexChannel", t))
    formLayout.addRow("Index channel", indexComboBox)

    rmComboBox = QtGui.QComboBox(leftWidget)
    rmComboBox.addItems(rmNames)
    if settings["ReferenceMaterial"] in rmNames:
        rmComboBox.setCurrentText(settings["ReferenceMaterial"])
    else:
        rmComboBox.setCurrentText(rmNames[0])
        drs.setSetting("ReferenceMaterial", rmNames[0])
    rmComboBox.currentTextChanged.connect(lambda t: drs.setSetting("ReferenceMaterial", t))
    formLayout.addRow("Reference material", rmComboBox)

    #Add checkbox to choose to use 85Rb to calculate 87Rb
    useRb85CheckBox = QtGui.QCheckBox(leftWidget)
    useRb85CheckBox.setChecked(settings["UseRb85"])
    useRb85CheckBox.toggled.connect(lambda t: drs.setSetting("UseRb85", bool(t)))
    formLayout.addRow("Use Rb85", useRb85CheckBox)

    formLayout.addRow(QtGui.QLabel(""))  # Spacer

    # --- Mask Settings ---
    maskLabel = QtGui.QLabel("<b>Mask Settings</b>")
    formLayout.addRow(maskLabel)

    maskCheckBox = QtGui.QCheckBox(leftWidget)
    maskCheckBox.setChecked(settings["Mask"])
    maskCheckBox.toggled.connect(lambda t: drs.setSetting("Mask", bool(t)))
    formLayout.addRow("Mask", maskCheckBox)

    maskComboBox = QtGui.QComboBox(leftWidget)
    maskComboBox.addItems(data.timeSeriesNames(data.Input))
    maskComboBox.setCurrentText(settings["MaskChannel"])
    maskComboBox.currentTextChanged.connect(
        lambda t: drs.setSetting("MaskChannel", t))
    formLayout.addRow("Mask channel", maskComboBox)

    maskLineEdit = QtGui.QLineEdit(leftWidget)
    maskLineEdit.setText(settings["MaskCutoff"])
    maskLineEdit.textChanged.connect(
        lambda t: drs.setSetting("MaskCutoff", float(t)))
    formLayout.addRow("Mask cutoff", maskLineEdit)

    maskTrimLineEdit = QtGui.QLineEdit(leftWidget)
    maskTrimLineEdit.setText(settings["MaskTrim"])
    maskTrimLineEdit.textChanged.connect(
        lambda t: drs.setSetting("MaskTrim", float(t)))
    formLayout.addRow("Mask trim", maskTrimLineEdit)

    formLayout.addRow(QtGui.QLabel(""))  # Spacer

    # --- Secondary RM Settings ---
    secondaryRMLabel = QtGui.QLabel("<b>Secondary RM Settings</b>")
    formLayout.addRow(secondaryRMLabel)

    #Secondary RM combo box
    secondaryComboBox = QtGui.QComboBox(leftWidget)
    secondaryComboBox.addItems(rmNames)

    # Set the current selection based on saved settings
    if settings["SecondaryRM"] in rmNames:
        secondaryComboBox.setCurrentText(settings["SecondaryRM"])
    elif rmNames:
        secondaryComboBox.setCurrentText(rmNames[0])
        drs.setSetting("SecondaryRM", rmNames[0])

    formLayout.addRow("Secondary RM", secondaryComboBox)

    # Regression model dropdown
    regressionModelComboBox = QtGui.QComboBox(leftWidget)
    regressionModelComboBox.addItems([
        "Model 1: York (MSWD scaling)",
        "Model 2: Total Least Squares",
        "Model 3: York + Overdispersion"
    ])
    current_model = settings.get("RegressionModel", 0)
    regressionModelComboBox.setCurrentIndex(int(current_model))
    regressionModelComboBox.currentIndexChanged.connect(lambda idx: drs.setSetting("RegressionModel", idx))
    formLayout.addRow("Regression Model", regressionModelComboBox)

    # Checkbox for using individual integrations for age correction
    useIndividualIntegrationsCheckBox = QtGui.QCheckBox(leftWidget)
    useIndividualIntegrationsCheckBox.setChecked(settings.get("UseIndividualIntegrations", False))
    useIndividualIntegrationsCheckBox.toggled.connect(lambda t: drs.setSetting("UseIndividualIntegrations", bool(t)))
    formLayout.addRow("Use Individual Integrations", useIndividualIntegrationsCheckBox)

    # Checkbox for using fixed intercept
    useFixedInterceptCheckBox = QtGui.QCheckBox(leftWidget)
    useFixedInterceptCheckBox.setChecked(settings.get("UseFixedIntercept", True))
    useFixedInterceptCheckBox.toggled.connect(lambda t: drs.setSetting("UseFixedIntercept", bool(t)))
    formLayout.addRow("Use Fixed Intercept", useFixedInterceptCheckBox)

    # Checkbox for using correlated errors
    useCorrelatedErrorsCheckBox = QtGui.QCheckBox(leftWidget)
    useCorrelatedErrorsCheckBox.setChecked(settings.get("UseCorrelatedErrors", False))
    useCorrelatedErrorsCheckBox.toggled.connect(lambda t: drs.setSetting("UseCorrelatedErrors", bool(t)))
    formLayout.addRow("Use Correlated Errors", useCorrelatedErrorsCheckBox)

    # Fixed intercept input
    fixedInterceptLineEdit = QtGui.QLineEdit(leftWidget)
    default_intercept = 0.7043
    current_index = secondaryComboBox.currentIndex
    if current_index >= 0:
        secondary_rm_name = secondaryComboBox.itemText(current_index)
        try:
            rm_data = data.referenceMaterialData(secondary_rm_name)
            if "87Sr/86Sr initial" in rm_data:
                default_intercept = float(rm_data["87Sr/86Sr initial"].value())
        except:
            pass
    intercept_value = float(settings.get("FixedIntercept", default_intercept))
    fixedInterceptLineEdit.setText(str(intercept_value))
    drs.setSetting("FixedIntercept", intercept_value)
    fixedInterceptLineEdit.textChanged.connect(
        lambda t: drs.setSetting("FixedIntercept", float(t) if t else 0.0))
    formLayout.addRow("Fixed Intercept (87Sr/86Sr)", fixedInterceptLineEdit)

    # Fixed intercept uncertainty input (2SE)
    fixedInterceptUncertLineEdit = QtGui.QLineEdit(leftWidget)
    default_intercept_uncert = 0.0002
    current_index = secondaryComboBox.currentIndex
    if current_index >= 0:
        secondary_rm_name = secondaryComboBox.itemText(current_index)
        try:
            rm_data = data.referenceMaterialData(secondary_rm_name)
            if "87Sr/86Sr initial" in rm_data:
                default_intercept_uncert = float(rm_data["87Sr/86Sr initial"].uncertainty())
        except:
            pass
    intercept_uncert_value = float(settings.get("FixedInterceptUncert", default_intercept_uncert))
    fixedInterceptUncertLineEdit.setText(str(intercept_uncert_value))
    drs.setSetting("FixedInterceptUncert", intercept_uncert_value)
    fixedInterceptUncertLineEdit.textChanged.connect(
        lambda t: drs.setSetting("FixedInterceptUncert", float(t) if t else 0.0))
    formLayout.addRow("Fixed Intercept Uncert (2SE)", fixedInterceptUncertLineEdit)

    # Secondary RM expected age input
    secondaryRMAgeLineEdit = QtGui.QLineEdit(leftWidget)
    default_age = 405
    current_index = secondaryComboBox.currentIndex
    if current_index >= 0:
        secondary_rm_name = secondaryComboBox.itemText(current_index)
        try:
            rm_data = data.referenceMaterialData(secondary_rm_name)
            if "Age" in rm_data:
                default_age = float(rm_data["Age"].value())
        except:
            pass

    age_value = float(settings.get("SecondaryRMAge", default_age))
    secondaryRMAgeLineEdit.setText(str(age_value))
    drs.setSetting("SecondaryRMAge", age_value)
    secondaryRMAgeLineEdit.textChanged.connect(
        lambda t: drs.setSetting("SecondaryRMAge", float(t) if t else 0.0))
    formLayout.addRow("Expected Age (Ma)", secondaryRMAgeLineEdit)

    # Secondary RM expected age uncertainty input
    secondaryRMAge2SELineEdit = QtGui.QLineEdit(leftWidget)
    default_age_2se = 0.0
    current_index = secondaryComboBox.currentIndex
    if current_index >= 0:
        secondary_rm_name = secondaryComboBox.itemText(current_index)
        try:
            rm_data = data.referenceMaterialData(secondary_rm_name)
            if "Age" in rm_data:
                default_age_2se = float(rm_data["Age"].uncertainty())
        except:
            pass
    age_2se_value = float(settings.get("SecondaryRMAge2SE", default_age_2se))
    secondaryRMAge2SELineEdit.setText(str(age_2se_value))
    drs.setSetting("SecondaryRMAge2SE", age_2se_value)
    secondaryRMAge2SELineEdit.textChanged.connect(
        lambda t: drs.setSetting("SecondaryRMAge2SE", float(t) if t else 0.0))
    formLayout.addRow("Age Uncertainty (Ma, 2SE)", secondaryRMAge2SELineEdit)

    # Connect the combo box to update both the setting AND the age line edit
    secondaryComboBox.currentTextChanged.connect(
        lambda t: [drs.setSetting("SecondaryRM", t),
                   updateSecondaryRMAge(t, secondaryRMAgeLineEdit, secondaryRMAge2SELineEdit,
                                       fixedInterceptLineEdit, fixedInterceptUncertLineEdit)])

    # Add spacer and reminder text
    formLayout.addRow(QtGui.QLabel(""))  # Spacer
    reminderLabel = QtGui.QLabel("<b>*** Refresh Secondary RM Results After Crunching ***</b>")
    formLayout.addRow(reminderLabel)

    # Add stretch to push everything up
    formLayout.addItem(QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Expanding))

    # Add left widget to main layout
    mainLayout.addWidget(leftWidget)

    # === RIGHT SIDE: Results Table ===
    rightWidget = QtGui.QWidget()
    rightLayout = QtGui.QVBoxLayout()
    rightWidget.setLayout(rightLayout)

    # Header row with label and buttons
    headerWidget = QtGui.QWidget(rightWidget)
    headerLayout = QtGui.QHBoxLayout(headerWidget)
    headerLayout.setContentsMargins(0, 0, 0, 0)

    # Results label on the left
    resultsLabel = QtGui.QLabel("<b>Secondary RM Results</b>")
    headerLayout.addWidget(resultsLabel)

    headerLayout.addStretch()  # Push buttons to the right

    # Refresh button
    refreshButton = QtGui.QPushButton("Refresh SecondaryRM Results")
    refreshButton.setToolTip("Reload results from the last DRS run")
    headerLayout.addWidget(refreshButton)

    # Export button
    exportButton = QtGui.QPushButton("Export to Excel")
    exportButton.setToolTip("Export results to an Excel file")
    headerLayout.addWidget(exportButton)

    rightLayout.addWidget(headerWidget)

    # Create the results table view
    resultsTableView = QtGui.QTableView(rightWidget)
    resultsTableModel = createResultsTableModel()
    resultsTableView.setModel(resultsTableModel)

    # Configure table appearance
    resultsTableView.setAlternatingRowColors(True)
    resultsTableView.setSelectionBehavior(QtGui.QAbstractItemView.SelectRows)
    resultsTableView.setSelectionMode(QtGui.QAbstractItemView.SingleSelection)
    resultsTableView.verticalHeader().setVisible(False)
    resultsTableView.horizontalHeader().setStretchLastSection(True)

    # Set column widths
    resultsTableView.setColumnWidth(0, 150)  # Parameter
    resultsTableView.setColumnWidth(1, 120)  # Value
    resultsTableView.setColumnWidth(2, 100)  # Uncertainty

    # Load any existing results
    try:
        current_settings = drs.settings()
        saved_results = current_settings.get("LastRegressionResults", None)
        if isinstance(saved_results, dict) and len(saved_results) > 0:
            updateResultsTableModel(resultsTableModel, saved_results)
    except Exception as e:
        print(f"Error loading saved results: {e}")

    rightLayout.addWidget(resultsTableView)

    # Connect button actions (defined after table model exists)
    def refreshResults():
        try:
            current_settings = drs.settings()
            saved = current_settings.get("LastRegressionResults", None)
            if isinstance(saved, dict) and len(saved) > 0:
                updateResultsTableModel(resultsTableModel, saved)
            else:
                updateResultsTableModel(resultsTableModel, None)
                QMessageBox.information(None, "No Results", "No results available yet.\n\nRun the DRS first to generate results.")
        except Exception as e:
            QMessageBox.warning(None, "Error", f"Error refreshing results: {e}")
    refreshButton.clicked.connect(refreshResults)

    def doExport():
        try:
            current_settings = drs.settings()
            saved = current_settings.get("LastRegressionResults", None)
            if isinstance(saved, dict):
                exportResults(saved)
            else:
                QMessageBox.warning(None, "Export Error", "No results to export. Run the DRS first.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.warning(None, "Export Error", f"Error exporting: {e}")
    exportButton.clicked.connect(doExport)

    # Add right widget to main layout
    mainLayout.addWidget(rightWidget)

    # Set stretch factors (equal width for left and right)
    mainLayout.setStretch(0, 1)  # Left (settings)
    mainLayout.setStretch(1, 1)  # Right (results table)

    drs.setSettingsWidget(widget)
